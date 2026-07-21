#!/usr/bin/env python3
"""Run the independent project audit against the V5 realistic demand fixture."""

from __future__ import annotations

import json
import statistics
import sys
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import audit_project_v4 as audit4


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data" / "Demand_Genie_Synthetic_Portfolio_v5.xlsx"
FORECAST_DIR = ROOT / "data" / "forecast-v5"
SPEND_DIR = ROOT / "data" / "spend-analysis-v5"
OUTPUT_PATH = ROOT / "data" / "audit-v5.json"
EXPECTED_DASHBOARD_HASHES = {
    "index.html": "ece5a062b805e19083ccb3ebcd0cac912b25cdc2f54e734c81e3eab223b808a9",
    "dashboard-v2.html": "24d843bd2bd2c31634d95024ea0d72973af094e9d08d9f3c6af95def9fca5b4e",
    "dashboard-v3.html": "7875d71c9d7f5e740e5cfd8b99f9ba4107a6fd520f15102c56c1cbd1ea82d035",
    "dashboard-v4.html": "e51bd79f26b37b1be585f4bbb98b907ae9785f4f44112e757418b41dac286130",
}
EXPECTED_PATTERNS = {"Smooth": 30, "Erratic": 18, "Intermittent": 17, "Lumpy": 15}


def observed_pattern(values: list[float]) -> tuple[float, float, str]:
    nonzero = [value for value in values if value > 0]
    adi = len(values) / len(nonzero)
    mean = statistics.fmean(nonzero)
    cv2 = (statistics.pstdev(nonzero) / mean) ** 2
    if adi < 1.32 and cv2 < 0.49:
        pattern = "Smooth"
    elif adi < 1.32:
        pattern = "Erratic"
    elif cv2 < 0.49:
        pattern = "Intermittent"
    else:
        pattern = "Lumpy"
    return adi, cv2, pattern


def audit_pattern_design(audit: audit4.Audit, workbook, histories: dict[str, list[tuple[str, float]]]) -> None:
    design = {str(row["SKU"]): row for row in audit4.workbook_records(workbook, "Demand_Pattern_Design")}
    mismatches = []
    distribution: Counter[str] = Counter()
    adi_values = []
    cv2_values = []
    for sku, history in histories.items():
        values = [value for _, value in sorted(history)]
        adi, cv2, pattern = observed_pattern(values)
        distribution[pattern] += 1
        adi_values.append(adi)
        cv2_values.append(cv2)
        row = design.get(sku)
        if (
            row is None
            or str(row["Target_Demand_Pattern"]) != pattern
            or str(row["Observed_Demand_Pattern"]) != pattern
            or not audit4.is_close(row["ADI"], adi)
            or not audit4.is_close(row["CV2_Nonzero_Demand"], cv2)
        ):
            mismatches.append(sku)
    audit.check(
        "V5 demand-pattern fixture matches declared ADI/CV2 targets",
        not mismatches and dict(distribution) == EXPECTED_PATTERNS and len(design) == 80,
        "30 smooth, 18 erratic, 17 intermittent, and 15 lumpy SKUs independently reproduced" if not mismatches else f"Mismatches: {mismatches[:5]}",
    )
    audit.metrics["demand_patterns"] = {
        "distribution": dict(sorted(distribution.items())),
        "adi_range": [round(min(adi_values), 4), round(max(adi_values), 4)],
        "cv2_range": [round(min(cv2_values), 4), round(max(cv2_values), 4)],
        "target_mismatches": len(mismatches),
    }


def main() -> int:
    audit4.WORKBOOK_PATH = WORKBOOK_PATH
    audit4.FORECAST_DIR = FORECAST_DIR
    audit4.SPEND_DIR = SPEND_DIR
    audit4.OUTPUT_PATH = OUTPUT_PATH
    audit4.EXPECTED_DASHBOARD_HASHES = EXPECTED_DASHBOARD_HASHES

    audit = audit4.Audit()
    audit4.audit_immutable_dashboards(audit)
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    required_sheets = {
        "Demand_Data", "Demand_History", "Demand_Pattern_Design", "Product_Master", "DDMRP_Positions",
        "DDMRP_Recommendations", "DDMRP_Master", "Inventory_Snapshot", "Open_Supply", "Qualified_Demand",
        "Spend_Lines", "Spend_Control", "Category_Risk", "Contracts",
    }
    missing = sorted(required_sheets - set(workbook.sheetnames))
    audit.check("Workbook contract contains all required sheets", not missing, "All required sheets present" if not missing else f"Missing: {', '.join(missing)}")
    if missing:
        result = {"status": "FAIL", "checks": audit.checks, "warnings": audit.warnings, "metrics": audit.metrics}
        OUTPUT_PATH.write_text(json.dumps(result, indent=2), encoding="utf-8")
        return 1

    sheets = {name: audit4.workbook_records(workbook, name) for name in required_sheets}
    histories = audit4.audit_demand(audit, sheets)
    audit_pattern_design(audit, workbook, histories)
    audit4.audit_ddmrp(audit, sheets, histories)
    audit4.audit_spend(audit, sheets)
    audit4.audit_item_segmentation(audit, sheets)
    audit4.audit_forecasts(audit, histories)
    audit4.audit_decomposition(audit)
    audit4.audit_manifest(audit)

    result: dict[str, Any] = {
        "status": "FAIL" if audit.failures else "PASS_WITH_WARNINGS" if audit.warnings else "PASS",
        "generated_at_utc": datetime.now(UTC).isoformat(),
        "checks": audit.checks,
        "warnings": audit.warnings,
        "metrics": audit.metrics,
    }
    OUTPUT_PATH.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(json.dumps(result, indent=2))
    return 1 if audit.failures else 0


if __name__ == "__main__":
    sys.exit(main())
