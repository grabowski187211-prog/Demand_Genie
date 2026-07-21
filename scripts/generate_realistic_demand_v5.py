#!/usr/bin/env python3
"""Create a V5 synthetic workbook with realistic ADI/CV2 demand-pattern diversity."""

from __future__ import annotations

import json
import math
import random
import statistics
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import generate_synthetic_demand_data as base


ROOT = Path(__file__).resolve().parents[1]
BASE_WORKBOOK = ROOT / "data" / "Demand_Genie_Synthetic_Demand_History.xlsx"
OUTPUT_WORKBOOK = ROOT / "data" / "Demand_Genie_Synthetic_Portfolio_v5.xlsx"

BEHAVIOR_PLAN = {
    "BRG": ("Smooth", "Smooth", "Smooth", "Erratic", "Smooth", "Erratic", "Intermittent", "Lumpy"),
    "GBX": ("Smooth", "Smooth", "Erratic", "Erratic", "Smooth", "Intermittent", "Intermittent", "Lumpy"),
    "HYP": ("Smooth", "Erratic", "Smooth", "Erratic", "Intermittent", "Lumpy", "Intermittent", "Lumpy"),
    "SEA": ("Smooth", "Smooth", "Smooth", "Smooth", "Smooth", "Erratic", "Intermittent", "Lumpy"),
    "CVL": ("Intermittent", "Lumpy", "Smooth", "Erratic", "Smooth", "Erratic", "Intermittent", "Lumpy"),
    "DRV": ("Smooth", "Smooth", "Erratic", "Erratic", "Smooth", "Intermittent", "Intermittent", "Lumpy"),
    "SNS": ("Smooth", "Erratic", "Smooth", "Erratic", "Smooth", "Intermittent", "Intermittent", "Lumpy"),
    "CNV": ("Smooth", "Smooth", "Erratic", "Smooth", "Erratic", "Intermittent", "Intermittent", "Lumpy"),
    "FST": ("Smooth", "Smooth", "Smooth", "Smooth", "Smooth", "Erratic", "Intermittent", "Lumpy"),
    "PCK": ("Intermittent", "Lumpy", "Erratic", "Lumpy", "Intermittent", "Lumpy", "Erratic", "Lumpy"),
}

LOW_VARIABILITY_PATTERNS = {"Smooth", "Intermittent"}
SPARSE_PATTERNS = {"Intermittent", "Lumpy"}
ERRATIC_FACTORS = (0.10, 0.18, 0.32, 0.55, 0.85, 1.25, 2.50, 4.10)


def headers(sheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in sheet[1]}


def event_months(count: int, offset: int, periods: int = 36) -> set[int]:
    """Distribute a declared number of demand events across the history."""
    return {int(math.floor((index + 0.5) * periods / count) + offset) % periods for index in range(count)}


def pattern_metrics(values: list[int]) -> dict[str, float | str | int]:
    nonzero = [value for value in values if value > 0]
    adi = len(values) / len(nonzero) if nonzero else math.inf
    nonzero_mean = statistics.fmean(nonzero) if nonzero else 0.0
    cv2 = (statistics.pstdev(nonzero) / nonzero_mean) ** 2 if nonzero_mean else math.inf
    overall_mean = statistics.fmean(values)
    demand_cv = statistics.pstdev(values) / overall_mean if overall_mean else math.inf
    if not nonzero:
        observed = "No demand"
    elif adi < 1.32 and cv2 < 0.49:
        observed = "Smooth"
    elif adi < 1.32:
        observed = "Erratic"
    elif cv2 < 0.49:
        observed = "Intermittent"
    else:
        observed = "Lumpy"
    xyz = "X" if demand_cv < 0.5 else "Y" if demand_cv < 1.0 else "Z"
    return {
        "nonzero_months": len(nonzero),
        "zero_share": 1 - len(nonzero) / len(values),
        "adi": adi,
        "cv2": cv2,
        "demand_cv": demand_cv,
        "observed": observed,
        "xyz": xyz,
    }


def stabilize_low_variability(values: list[int]) -> list[int]:
    nonzero = [value for value in values if value > 0]
    center = statistics.fmean(nonzero)
    return [0 if value == 0 else max(1, int(round(center * 0.55 + value * 0.45))) for value in values]


def amplify_nonzero_variability(values: list[int]) -> list[int]:
    nonzero_positions = [index for index, value in enumerate(values) if value > 0]
    center = statistics.fmean(values[index] for index in nonzero_positions)
    normalized_mean = statistics.fmean(ERRATIC_FACTORS)
    output = list(values)
    for event_index, position in enumerate(nonzero_positions):
        factor = ERRATIC_FACTORS[event_index % len(ERRATIC_FACTORS)] / normalized_mean
        output[position] = max(1, int(round(center * factor)))
    return output


def generate_history(group: dict[str, Any], sku_index: int, group_index: int, target: str) -> list[int]:
    seed = base.SEED + 50_000 + group_index * 1_000 + sku_index * 37
    rng = random.Random(seed)
    if target == "Smooth":
        event_count = 36 - ((sku_index + group_index) % 3)
    elif target == "Erratic":
        event_count = 36 - ((sku_index + 2 * group_index) % 6)
    elif target == "Intermittent":
        event_count = (24, 22, 20, 18, 16)[(sku_index + group_index) % 5]
    else:
        event_count = (22, 19, 16, 13, 10)[(sku_index + group_index) % 5]
    active_months = event_months(event_count, offset=(group_index * 3 + sku_index * 5) % 36)

    values: list[int] = []
    event_index = 0
    for month_index in range(36):
        if month_index not in active_months:
            values.append(0)
            continue
        level = float(group["base"]) * base.SKU_MULTIPLIERS[sku_index]
        phase = sku_index * 0.61 + group_index * 0.27
        seasonal = 1 + float(group["seasonality"]) * math.sin(month_index / 12 * 2 * math.pi + phase)
        trend = 1 + float(group["trend"]) * month_index
        if target in LOW_VARIABILITY_PATTERNS:
            amplitude = max(0.65, 1 + rng.gauss(0, min(float(group["volatility"]), 0.12)))
        else:
            amplitude = ERRATIC_FACTORS[(event_index + sku_index + group_index) % len(ERRATIC_FACTORS)]
            amplitude /= statistics.fmean(ERRATIC_FACTORS)
            amplitude *= rng.uniform(0.92, 1.08)
        if str(group["code"]) == "SNS" and month_index >= 24:
            trend *= 1.10
        values.append(max(1, int(round(level * seasonal * trend * amplitude))))
        event_index += 1

    for _ in range(4):
        metrics = pattern_metrics(values)
        if metrics["observed"] == target:
            return values
        values = stabilize_low_variability(values) if target in LOW_VARIABILITY_PATTERNS else amplify_nonzero_variability(values)
    metrics = pattern_metrics(values)
    if metrics["observed"] != target:
        raise AssertionError(f"Could not create {target} history for {group['code']}-{sku_index + 1:02d}: {metrics}")
    return values


def replace_rows(sheet, rows: list[list[object]]) -> None:
    style_templates = [copy(cell._style) for cell in sheet[2]] if sheet.max_row >= 2 else []
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for row in rows:
        sheet.append(row)
    if style_templates:
        for row in sheet.iter_rows(min_row=2):
            for index, cell in enumerate(row):
                if index < len(style_templates):
                    cell._style = copy(style_templates[index])
    last_cell = f"{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for table in sheet.tables.values():
        table.ref = f"A1:{last_cell}"


def assign_abc(histories: dict[str, list[int]], products: dict[str, dict[str, Any]]) -> dict[str, str]:
    usage_values = {
        sku: statistics.fmean(history) * 12 * float(products[sku]["Unit_Cost_EUR"])
        for sku, history in histories.items()
    }
    total = sum(usage_values.values())
    cumulative = 0.0
    output: dict[str, str] = {}
    for sku, value in sorted(usage_values.items(), key=lambda item: (-item[1], item[0])):
        prior_share = cumulative / total if total else 0.0
        output[sku] = "A" if prior_share < 0.8 else "B" if prior_share < 0.95 else "C"
        cumulative += value
    return output


def update_read_me(workbook) -> None:
    sheet = workbook["Read_Me"]
    replacements = {
        "Purpose": "Synthetic monthly demand with validated smooth, erratic, intermittent, and lumpy item behavior; complete DDMRP planning data; and an auditable procurement spend ledger. All values are synthetic demo data.",
        "Data note": "V5 contains explicit zero-demand periods and varied nonzero demand sizes. Every SKU is validated against the declared ADI 1.32 and CV2 0.49 behavior boundaries.",
        "Portfolio method": "Demand_History is the separate consumption fact for ABC/XYZ and ADI/CV2 behavior. Demand_Pattern_Design records the intended and observed synthetic behavior for transparent fixture validation.",
    }
    for row in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row, 1).value or "")
        if label in replacements:
            sheet.cell(row, 2).value = replacements[label]
    row = sheet.max_row + 1
    sheet.cell(row, 1).value = "Demand_Pattern_Design"
    sheet.cell(row, 2).value = "V5 synthetic target and observed ADI/CV2 evidence by SKU."
    sheet.cell(row, 1)._style = copy(sheet.cell(3, 1)._style)
    sheet.cell(row, 2)._style = copy(sheet.cell(3, 2)._style)


def main() -> None:
    workbook = load_workbook(BASE_WORKBOOK)
    group_lookup = {str(group["code"]): group for group in base.GROUPS}
    group_order = {str(group["code"]): index for index, group in enumerate(base.GROUPS)}
    product_sheet = workbook["Product_Master"]
    product_headers = headers(product_sheet)
    products: dict[str, dict[str, Any]] = {}
    product_rows: dict[str, int] = {}
    for row_index in range(2, product_sheet.max_row + 1):
        row = {name: product_sheet.cell(row_index, column).value for name, column in product_headers.items()}
        sku = str(row["SKU"])
        products[sku] = row
        product_rows[sku] = row_index

    histories: dict[str, list[int]] = {}
    targets: dict[str, str] = {}
    metrics_by_sku: dict[str, dict[str, float | str | int]] = {}
    for sku, product in products.items():
        group_code = str(product["Product_Group_Code"])
        sku_index = int(sku.rsplit("-", 1)[1]) - 1
        target = BEHAVIOR_PLAN[group_code][sku_index]
        history = generate_history(group_lookup[group_code], sku_index, group_order[group_code], target)
        metrics = pattern_metrics(history)
        if metrics["observed"] != target:
            raise AssertionError(f"Pattern validation failed for {sku}: target {target}, observed {metrics['observed']}")
        histories[sku] = history
        targets[sku] = target
        metrics_by_sku[sku] = metrics

    abc_by_sku = assign_abc(histories, products)
    for sku, row_index in product_rows.items():
        product_sheet.cell(row_index, product_headers["Demand_Profile"]).value = targets[sku]
        product_sheet.cell(row_index, product_headers["ABC_Class"]).value = abc_by_sku[sku]
        product_sheet.cell(row_index, product_headers["XYZ_Class"]).value = metrics_by_sku[sku]["xyz"]

    demand_sheet = workbook["Demand_Data"]
    demand_headers = headers(demand_sheet)
    month_index = {month.strftime("%Y-%m"): index for index, month in enumerate(base.MONTHS)}
    for row_index in range(2, demand_sheet.max_row + 1):
        sku = str(demand_sheet.cell(row_index, demand_headers["SKU"]).value)
        month = demand_sheet.cell(row_index, demand_headers["Month"]).value.strftime("%Y-%m")
        units = histories[sku][month_index[month]]
        price = float(demand_sheet.cell(row_index, demand_headers["Unit_Price_EUR"]).value)
        demand_sheet.cell(row_index, demand_headers["Demand_Units"]).value = units
        demand_sheet.cell(row_index, demand_headers["Demand_Value_EUR"]).value = round(units * price, 2)
        demand_sheet.cell(row_index, demand_headers["Demand_Profile"]).value = targets[sku]

    matrix_sheet = workbook["Demand_Matrix"]
    matrix_headers = headers(matrix_sheet)
    for row_index in range(2, matrix_sheet.max_row + 1):
        sku = str(matrix_sheet.cell(row_index, matrix_headers["SKU"]).value)
        matrix_sheet.cell(row_index, matrix_headers["Demand_Profile"]).value = targets[sku]
        for month, value in zip(base.MONTHS, histories[sku]):
            matrix_sheet.cell(row_index, matrix_headers[month.strftime("%Y-%m")]).value = value

    demand_history_sheet = workbook["Demand_History"]
    history_headers = headers(demand_history_sheet)
    for row_index in range(2, demand_history_sheet.max_row + 1):
        sku = str(demand_history_sheet.cell(row_index, history_headers["Part_ID"]).value)
        month = demand_history_sheet.cell(row_index, history_headers["Period"]).value.strftime("%Y-%m")
        demand_history_sheet.cell(row_index, history_headers["Demand_Units"]).value = histories[sku][month_index[month]]
        demand_history_sheet.cell(row_index, history_headers["Data_Quality_Status"]).value = "Complete V5 synthetic history with explicit zero periods"

    all_skus = []
    for sku in sorted(products, key=lambda value: (group_order[value.split("-")[0]], value)):
        product = products[sku]
        all_skus.append(
            {
                "sku": sku,
                "description": product["SKU_Description"],
                "product_group": product["Product_Group"],
                "product_group_code": product["Product_Group_Code"],
                "supply_type": product["Supply_Type"],
                "lead_time": product["Lead_Time_Days"],
                "xyz_class": metrics_by_sku[sku]["xyz"],
            }
        )
    ddmrp = base.build_ddmrp_rows(all_skus, histories, random.Random(base.SEED + 51_000))
    ddmrp_sheets = {
        "DDMRP_Decoupling": "decoupling",
        "DDMRP_Master": "master",
        "Inventory_Snapshot": "inventory",
        "Open_Supply": "supply",
        "Qualified_Demand": "qualified_demand",
        "DDMRP_Positions": "positions",
        "DDMRP_Recommendations": "recommendations",
        "DDMRP_Adjustments": "adjustments",
    }
    for sheet_name, row_key in ddmrp_sheets.items():
        replace_rows(workbook[sheet_name], ddmrp[row_key])

    if "Demand_Pattern_Design" in workbook.sheetnames:
        del workbook["Demand_Pattern_Design"]
    pattern_index = workbook.sheetnames.index("Demand_History") + 1
    pattern_sheet = workbook.create_sheet("Demand_Pattern_Design", pattern_index)
    pattern_sheet.append(
        [
            "SKU", "Product_Group", "Target_Demand_Pattern", "Nonzero_Months", "Zero_Share", "ADI",
            "CV2_Nonzero_Demand", "Observed_Demand_Pattern", "Demand_CV", "XYZ_Class",
            "Annual_Usage_Value_EUR", "Design_Note",
        ]
    )
    for sku in sorted(products):
        product = products[sku]
        metrics = metrics_by_sku[sku]
        annual_value = statistics.fmean(histories[sku]) * 12 * float(product["Unit_Cost_EUR"])
        pattern_sheet.append(
            [
                sku, product["Product_Group"], targets[sku], metrics["nonzero_months"], metrics["zero_share"],
                metrics["adi"], metrics["cv2"], metrics["observed"], metrics["demand_cv"], metrics["xyz"],
                annual_value, "Deterministic synthetic fixture; validate policy on real uncensored demand",
            ]
        )
    base.style_sheet(
        pattern_sheet,
        {"A": 14, "B": 24, "C": 23, "D": 16, "E": 13, "F": 12, "G": 23, "H": 25, "I": 14, "J": 12, "K": 24, "L": 68},
    )
    base.add_table(pattern_sheet, "DemandPatternDesignTable")
    base.format_columns(
        pattern_sheet,
        integers=("Nonzero_Months",),
        decimals=("Zero_Share", "ADI", "CV2_Nonzero_Demand", "Demand_CV", "Annual_Usage_Value_EUR"),
    )
    update_read_me(workbook)

    OUTPUT_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_WORKBOOK)
    distribution = Counter(str(metrics["observed"]) for metrics in metrics_by_sku.values())
    result = {
        "status": "PASS",
        "output": str(OUTPUT_WORKBOOK.relative_to(ROOT)),
        "sku_count": len(histories),
        "pattern_distribution": dict(sorted(distribution.items())),
        "adi_range": [round(min(float(metrics["adi"]) for metrics in metrics_by_sku.values()), 3), round(max(float(metrics["adi"]) for metrics in metrics_by_sku.values()), 3)],
        "cv2_range": [round(min(float(metrics["cv2"]) for metrics in metrics_by_sku.values()), 3), round(max(float(metrics["cv2"]) for metrics in metrics_by_sku.values()), 3)],
        "target_mismatches": sum(metrics_by_sku[sku]["observed"] != targets[sku] for sku in histories),
    }
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
