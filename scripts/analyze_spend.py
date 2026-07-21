#!/usr/bin/env python3
"""Create auditable procurement spend and optional inventory segmentation outputs."""

from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled with a clear runtime message
    load_workbook = None


SPEND_REQUIRED = {
    "Source_System",
    "Transaction_ID",
    "Line_ID",
    "Transaction_Date",
    "Supplier_ID",
    "Supplier_Name",
    "Category_L1",
    "Spend_Base",
    "Base_Currency",
}

DEMAND_REQUIRED = {
    "Period",
    "Part_ID",
    "Location",
    "Demand_Units",
    "Unit_Cost_Base",
    "Base_Currency",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Spend CSV or Excel workbook")
    parser.add_argument("output_dir", type=Path, help="Directory for analysis outputs")
    parser.add_argument("--spend-sheet", default="Spend_Lines")
    parser.add_argument("--demand-file", type=Path)
    parser.add_argument("--demand-sheet", default="Demand_History")
    parser.add_argument("--control-total", type=float)
    parser.add_argument("--abc-a", type=float, default=0.80, help="A cumulative share boundary")
    parser.add_argument("--abc-b", type=float, default=0.95, help="B cumulative share boundary")
    parser.add_argument("--xyz-x", type=float, default=0.50, help="Upper CV boundary for X")
    parser.add_argument("--xyz-y", type=float, default=1.00, help="Upper CV boundary for Y")
    parser.add_argument("--adi-boundary", type=float, default=1.32)
    parser.add_argument("--cv2-boundary", type=float, default=0.49)
    parser.add_argument("--periods-per-year", type=int, default=12)
    parser.add_argument("--kraljic-threshold", type=float, default=50.0)
    parser.add_argument("--fragmented-supplier-count", type=int, default=5)
    parser.add_argument("--concentration-share", type=float, default=0.80)
    args = parser.parse_args()
    if not 0 < args.abc_a < args.abc_b <= 1:
        parser.error("ABC boundaries must satisfy 0 < A < B <= 1")
    if not 0 <= args.xyz_x < args.xyz_y:
        parser.error("XYZ boundaries must satisfy 0 <= X < Y")
    if args.periods_per_year <= 0:
        parser.error("--periods-per-year must be positive")
    return args


def read_table(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as source:
            reader = csv.DictReader(source)
            headers = [str(value).strip() for value in (reader.fieldnames or [])]
            return headers, [{str(key).strip(): value for key, value in row.items()} for row in reader]
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported input format: {path.suffix}. Use CSV, XLSX, or XLSM.")
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for Excel input")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if not sheet_name or sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {path.name}")
    values = workbook[sheet_name].iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(values)]
    except StopIteration:
        return [], []
    rows = [dict(zip(headers, row)) for row in values if any(value is not None for value in row)]
    return headers, rows


def workbook_has_sheet(path: Path, sheet_name: str) -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"} or load_workbook is None:
        return False
    workbook = load_workbook(path, read_only=True, data_only=True)
    return sheet_name in workbook.sheetnames


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if math.isfinite(float(value)) else None
    try:
        parsed = float(str(value).strip())
        return parsed if math.isfinite(parsed) else None
    except (TypeError, ValueError):
        return None


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = text(value)
    if not raw:
        return None
    for candidate in (raw, raw[:10]):
        try:
            return datetime.fromisoformat(candidate.replace("Z", "+00:00")).date().isoformat()
        except ValueError:
            continue
    return None


def month_text(value: Any) -> str | None:
    parsed = date_text(value)
    return parsed[:7] if parsed else None


def continuous_months(first: str, last: str) -> list[str]:
    """Return every YYYY-MM period in the inclusive range."""
    first_year, first_month = (int(part) for part in first.split("-"))
    last_year, last_month = (int(part) for part in last.split("-"))
    start = first_year * 12 + first_month - 1
    stop = last_year * 12 + last_month - 1
    return [f"{index // 12:04d}-{index % 12 + 1:02d}" for index in range(start, stop + 1)]


def boolean(value: Any) -> bool | None:
    raw = text(value).lower()
    if raw in {"true", "yes", "y", "1", "on contract", "preferred"}:
        return True
    if raw in {"false", "no", "n", "0", "off contract", "non-preferred"}:
        return False
    return None


def require_columns(headers: Iterable[str], required: set[str], label: str) -> None:
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{label} is missing required columns: {', '.join(missing)}")


def write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str] | None = None) -> None:
    if headers is None:
        headers = list(rows[0]) if rows else []
    with path.open("w", newline="", encoding="utf-8") as target:
        writer = csv.DictWriter(target, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def safe_ratio(numerator: float, denominator: float) -> float | None:
    return numerator / denominator if denominator else None


def spend_parent(row: dict[str, Any]) -> tuple[str, str]:
    candidates = [
        ("Supplier_Parent_ID", "Supplier_Parent_Name"),
        ("Supplier_Normalized_ID", "Supplier_Normalized_Name"),
        ("Supplier_ID", "Supplier_Name"),
    ]
    for id_field, name_field in candidates:
        if text(row.get(id_field)):
            return text(row.get(id_field)), text(row.get(name_field)) or text(row.get("Supplier_Name"))
    return "UNMAPPED", text(row.get("Supplier_Name")) or "Unmapped supplier"


def assign_pareto(
    rows: list[dict[str, Any]],
    value_field: str,
    a_boundary: float,
    b_boundary: float,
    *,
    class_field: str = "Spend_ABC",
    share_field: str = "Positive_Spend_Share",
    cumulative_field: str = "Cumulative_Positive_Spend_Share",
    distance_field: str = "ABC_Boundary_Distance",
) -> list[dict[str, Any]]:
    ordered = sorted(rows, key=lambda row: (-max(0.0, float(row[value_field])), str(row)))
    total = sum(max(0.0, float(row[value_field])) for row in ordered)
    cumulative = 0.0
    for rank, row in enumerate(ordered, 1):
        value = max(0.0, float(row[value_field]))
        prior_share = cumulative / total if total else 0.0
        cumulative += value
        cumulative_share = cumulative / total if total else 0.0
        klass = "A" if prior_share < a_boundary else "B" if prior_share < b_boundary else "C"
        row.update(
            {
                "Rank": rank,
                share_field: safe_ratio(value, total),
                cumulative_field: cumulative_share,
                class_field: klass,
                distance_field: min(abs(cumulative_share - a_boundary), abs(cumulative_share - b_boundary)),
            }
        )
    return ordered


def normalize_spend(raw_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    normalized: list[dict[str, Any]] = []
    counters: defaultdict[str, int] = defaultdict(int)
    seen: defaultdict[tuple[str, str, str], int] = defaultdict(int)
    today = date.today().isoformat()
    for source_row in raw_rows:
        row = dict(source_row)
        spend = number(row.get("Spend_Base"))
        transaction_date = date_text(row.get("Transaction_Date"))
        if spend is None:
            counters["invalid_spend"] += 1
            continue
        if transaction_date is None:
            counters["invalid_date"] += 1
            continue
        row["Spend_Base"] = spend
        row["Transaction_Date"] = transaction_date
        row["Analysis_Month"] = transaction_date[:7]
        row["Positive_Spend"] = max(0.0, spend)
        row["Credit_Spend"] = min(0.0, spend)
        row["Supplier_Analysis_ID"], row["Supplier_Analysis_Name"] = spend_parent(row)
        row["On_Contract_Parsed"] = boolean(row.get("On_Contract"))
        row["Preferred_Supplier_Parsed"] = boolean(row.get("Preferred_Supplier"))
        key = (text(row.get("Source_System")), text(row.get("Transaction_ID")), text(row.get("Line_ID")))
        seen[key] += 1
        if transaction_date > today:
            counters["future_date"] += 1
        normalized.append(row)
    counters["duplicate_key_rows"] = sum(count for count in seen.values() if count > 1)
    counters["duplicate_keys"] = sum(1 for count in seen.values() if count > 1)
    return normalized, dict(counters)


def supplier_summary(rows: list[dict[str, Any]], args: argparse.Namespace) -> list[dict[str, Any]]:
    groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[row["Supplier_Analysis_ID"]].append(row)
    output = []
    for supplier_id, members in groups.items():
        output.append(
            {
                "Supplier_ID": supplier_id,
                "Supplier_Name": members[0]["Supplier_Analysis_Name"],
                "Net_Spend": sum(row["Spend_Base"] for row in members),
                "Positive_Spend": sum(row["Positive_Spend"] for row in members),
                "Credits": sum(row["Credit_Spend"] for row in members),
                "Transaction_Count": len({(text(row.get("Source_System")), text(row.get("Transaction_ID"))) for row in members}),
                "Line_Count": len(members),
                "Category_Count": len({text(row.get("Category_L1")) for row in members if text(row.get("Category_L1"))}),
                "Site_Count": len({text(row.get("Plant")) for row in members if text(row.get("Plant"))}),
            }
        )
    return assign_pareto(output, "Positive_Spend", args.abc_a, args.abc_b)


def kraljic_quadrant(impact: float | None, risk: float | None, threshold: float) -> str:
    if impact is None or risk is None:
        return "Unscored"
    if impact >= threshold and risk >= threshold:
        return "Strategic"
    if impact >= threshold:
        return "Leverage"
    if risk >= threshold:
        return "Bottleneck"
    return "Routine"


def weighted_score(rows: list[dict[str, Any]], field: str) -> tuple[float | None, float]:
    scored = [(number(row.get(field)), row["Positive_Spend"]) for row in rows]
    scored = [(score, weight) for score, weight in scored if score is not None and 0 <= score <= 100 and weight > 0]
    covered = sum(weight for _, weight in scored)
    if not covered:
        return None, 0.0
    return sum(score * weight for score, weight in scored) / covered, covered


def category_summary(rows: list[dict[str, Any]], args: argparse.Namespace) -> list[dict[str, Any]]:
    groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[text(row.get("Category_L1")) or "UNCLASSIFIED"].append(row)
    output = []
    for category, members in groups.items():
        positive = sum(row["Positive_Spend"] for row in members)
        supplier_spend: defaultdict[str, float] = defaultdict(float)
        for row in members:
            supplier_spend[row["Supplier_Analysis_ID"]] += row["Positive_Spend"]
        shares = sorted((value / positive for value in supplier_spend.values()), reverse=True) if positive else []
        impact, impact_covered = weighted_score(members, "Business_Impact_Score")
        risk, risk_covered = weighted_score(members, "Supply_Risk_Score")
        contract_known = sum(row["Positive_Spend"] for row in members if row["On_Contract_Parsed"] is not None)
        on_contract = sum(row["Positive_Spend"] for row in members if row["On_Contract_Parsed"] is True)
        output.append(
            {
                "Category_L1": category,
                "Net_Spend": sum(row["Spend_Base"] for row in members),
                "Positive_Spend": positive,
                "Credits": sum(row["Credit_Spend"] for row in members),
                "Supplier_Count": len([value for value in supplier_spend.values() if value > 0]),
                "Transaction_Count": len({(text(row.get("Source_System")), text(row.get("Transaction_ID"))) for row in members}),
                "Top_1_Supplier_Share": shares[0] if shares else None,
                "Top_3_Supplier_Share": sum(shares[:3]) if shares else None,
                "HHI_0_10000": sum((share * 100) ** 2 for share in shares),
                "Effective_Supplier_Count": 1 / sum(share**2 for share in shares) if shares else None,
                "Contract_Status_Coverage": safe_ratio(contract_known, positive),
                "On_Contract_Share_of_Known": safe_ratio(on_contract, contract_known),
                "Business_Impact_Score": impact,
                "Business_Impact_Score_Coverage": safe_ratio(impact_covered, positive),
                "Supply_Risk_Score": risk,
                "Supply_Risk_Score_Coverage": safe_ratio(risk_covered, positive),
                "Kraljic_Quadrant": kraljic_quadrant(impact, risk, args.kraljic_threshold),
            }
        )
    return assign_pareto(output, "Positive_Spend", args.abc_a, args.abc_b)


def monthly_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[row["Analysis_Month"]].append(row)
    return [
        {
            "Month": month,
            "Net_Spend": sum(row["Spend_Base"] for row in members),
            "Positive_Spend": sum(row["Positive_Spend"] for row in members),
            "Credits": sum(row["Credit_Spend"] for row in members),
            "Supplier_Count": len({row["Supplier_Analysis_ID"] for row in members}),
            "Transaction_Count": len({(text(row.get("Source_System")), text(row.get("Transaction_ID"))) for row in members}),
        }
        for month, members in sorted(groups.items())
    ]


def spend_coverage(rows: list[dict[str, Any]], predicate) -> float | None:
    total = sum(row["Positive_Spend"] for row in rows)
    covered = sum(row["Positive_Spend"] for row in rows if predicate(row))
    return safe_ratio(covered, total)


def quality_row(check: str, value: Any, status: str, detail: str) -> dict[str, Any]:
    return {"Check": check, "Value": value, "Status": status, "Detail": detail}


def spend_quality(raw_count: int, rows: list[dict[str, Any]], counters: dict[str, int], args: argparse.Namespace) -> list[dict[str, Any]]:
    currencies = sorted({text(row.get("Base_Currency")) for row in rows if text(row.get("Base_Currency"))})
    analyzed = sum(row["Spend_Base"] for row in rows)
    difference = analyzed - args.control_total if args.control_total is not None else None
    tolerance = 0.01
    key_gaps = sum(
        1
        for row in rows
        if not all(text(row.get(field)) for field in ("Source_System", "Transaction_ID", "Line_ID"))
    )
    category_coverage = spend_coverage(rows, lambda row: bool(text(row.get("Category_L1"))))
    raw_supplier_coverage = spend_coverage(rows, lambda row: bool(text(row.get("Supplier_ID"))))
    output = [
        quality_row("Input rows", raw_count, "INFO", "Rows read before validation"),
        quality_row("Analyzed rows", len(rows), "PASS" if len(rows) == raw_count else "WARN", "Rows with valid spend and date"),
        quality_row("Invalid spend rows", counters.get("invalid_spend", 0), "PASS" if not counters.get("invalid_spend") else "FAIL", "Excluded from calculations; correct at source"),
        quality_row("Invalid date rows", counters.get("invalid_date", 0), "PASS" if not counters.get("invalid_date") else "FAIL", "Excluded from calculations; correct at source"),
        quality_row("Duplicate transaction-line keys", counters.get("duplicate_keys", 0), "PASS" if not counters.get("duplicate_keys") else "FAIL", "Duplicates remain in totals; resolve before decisions"),
        quality_row("Missing transaction-line keys", key_gaps, "PASS" if not key_gaps else "FAIL", "Source, transaction, and line IDs are required for lineage"),
        quality_row("Future-dated rows", counters.get("future_date", 0), "PASS" if not counters.get("future_date") else "WARN", "Validate scope and posting dates"),
        quality_row("Base currencies", ", ".join(currencies) or "Missing", "PASS" if len(currencies) == 1 else "FAIL", "Exactly one base currency is required"),
        quality_row("Raw supplier coverage by positive spend", raw_supplier_coverage, "PASS" if raw_supplier_coverage == 1 else "FAIL", "A source supplier is required for every spend line"),
        quality_row("Category coverage by positive spend", category_coverage, "PASS" if category_coverage == 1 else "WARN", "Unclassified spend must remain visible"),
        quality_row("Normalized supplier coverage by positive spend", spend_coverage(rows, lambda row: bool(text(row.get("Supplier_Normalized_ID")) or text(row.get("Supplier_Parent_ID")))), "INFO", "Raw supplier fallback is used when absent"),
        quality_row("Contract-status coverage by positive spend", spend_coverage(rows, lambda row: row["On_Contract_Parsed"] is not None), "INFO", "Unknown is not treated as off-contract"),
    ]
    if args.control_total is None:
        output.append(quality_row("Control-total reconciliation", "Not supplied", "WARN", "Provide --control-total before financial interpretation"))
    else:
        output.append(quality_row("Control-total reconciliation", difference, "PASS" if abs(difference) <= tolerance else "FAIL", f"Analyzed {analyzed}; control {args.control_total}; tolerance {tolerance}"))
    return output


def normalize_demand(raw_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    rows = []
    counters: defaultdict[str, int] = defaultdict(int)
    seen: defaultdict[tuple[str, str, str], int] = defaultdict(int)
    for source_row in raw_rows:
        row = dict(source_row)
        period = month_text(row.get("Period"))
        demand = number(row.get("Demand_Units"))
        cost = number(row.get("Unit_Cost_Base"))
        if period is None:
            counters["invalid_period"] += 1
            continue
        if demand is None or cost is None:
            counters["invalid_numeric"] += 1
            continue
        if demand < 0:
            counters["negative_demand"] += 1
        if cost <= 0:
            counters["nonpositive_cost"] += 1
        row["Period"] = period
        row["Demand_Units"] = demand
        row["Unit_Cost_Base"] = cost
        key = (text(row.get("Part_ID")), text(row.get("Location")), period)
        seen[key] += 1
        rows.append(row)
    counters["duplicate_keys"] = sum(1 for count in seen.values() if count > 1)
    return rows, dict(counters)


def xyz_class(cv: float | None, x_boundary: float, y_boundary: float) -> str:
    if cv is None:
        return "No demand"
    if cv < x_boundary:
        return "X"
    if cv < y_boundary:
        return "Y"
    return "Z"


def demand_pattern(adi: float | None, cv2: float | None, adi_boundary: float, cv2_boundary: float) -> str:
    if adi is None or cv2 is None:
        return "No demand"
    if adi < adi_boundary and cv2 < cv2_boundary:
        return "Smooth"
    if adi < adi_boundary:
        return "Erratic"
    if cv2 < cv2_boundary:
        return "Intermittent"
    return "Lumpy"


def item_segmentation(rows: list[dict[str, Any]], args: argparse.Namespace) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    observed_periods = sorted({row["Period"] for row in rows})
    expected_periods = continuous_months(observed_periods[0], observed_periods[-1]) if observed_periods else []
    groups: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[(text(row.get("Part_ID")), text(row.get("Location")))].append(row)
    output = []
    incomplete = 0
    for (part, location), members in groups.items():
        by_period: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in members:
            by_period[row["Period"]].append(row)
        complete = set(by_period) == set(expected_periods) and all(len(values) == 1 for values in by_period.values())
        if not complete:
            incomplete += 1
        ordered = [by_period[period][0] for period in expected_periods if len(by_period.get(period, [])) == 1]
        valid = complete and all(row["Demand_Units"] >= 0 and row["Unit_Cost_Base"] > 0 for row in ordered)
        values = [row["Demand_Units"] for row in ordered] if valid else []
        mean = statistics.fmean(values) if values else None
        cv = statistics.pstdev(values) / mean if values and mean and len(values) > 1 else (0.0 if mean else None)
        nonzero = [value for value in values if value > 0]
        adi = len(values) / len(nonzero) if nonzero else None
        nz_mean = statistics.fmean(nonzero) if nonzero else None
        cv2 = (statistics.pstdev(nonzero) / nz_mean) ** 2 if nonzero and nz_mean and len(nonzero) > 1 else (0.0 if nonzero else None)
        latest_cost = ordered[-1]["Unit_Cost_Base"] if ordered else None
        annual_units = statistics.fmean(values) * args.periods_per_year if values else None
        annual_value = annual_units * latest_cost if annual_units is not None and latest_cost is not None else 0.0
        xyz = xyz_class(cv, args.xyz_x, args.xyz_y) if valid else "Invalid history"
        pattern = demand_pattern(adi, cv2, args.adi_boundary, args.cv2_boundary) if valid else "Invalid history"
        output.append(
            {
                "Part_ID": part,
                "Location": location,
                "History_Start": expected_periods[0] if expected_periods else None,
                "History_End": expected_periods[-1] if expected_periods else None,
                "Period_Count": len(values),
                "Complete_History": "Yes" if complete else "No",
                "Mean_Demand": mean,
                "Zero_Share": safe_ratio(len(values) - len(nonzero), len(values)),
                "Demand_CV": cv,
                "XYZ": xyz,
                "XYZ_Boundary_Distance": min(abs(cv - args.xyz_x), abs(cv - args.xyz_y)) if cv is not None else None,
                "ADI": adi,
                "CV2_Nonzero_Demand": cv2,
                "Demand_Pattern": pattern,
                "ADI_Boundary_Distance": abs(adi - args.adi_boundary) if adi is not None else None,
                "CV2_Boundary_Distance": abs(cv2 - args.cv2_boundary) if cv2 is not None else None,
                "Approved_Unit_Cost_Base": latest_cost,
                "Annualized_Usage_Units": annual_units,
                "Annual_Usage_Value": annual_value,
                "Criticality": text(ordered[-1].get("Criticality")) if ordered else "",
                "Lifecycle_Status": text(ordered[-1].get("Lifecycle_Status")) if ordered else "",
            }
        )
    eligible = [row for row in output if row["Complete_History"] == "Yes" and row["Annual_Usage_Value"] > 0 and row["Demand_Pattern"] != "Invalid history"]
    excluded = [row for row in output if row not in eligible]
    ranked = assign_pareto(
        eligible,
        "Annual_Usage_Value",
        args.abc_a,
        args.abc_b,
        class_field="Usage_Value_ABC",
        share_field="Annual_Usage_Value_Share",
        cumulative_field="Cumulative_Annual_Usage_Value_Share",
        distance_field="Usage_Value_ABC_Boundary_Distance",
    )
    for row in ranked:
        row["Segmentation_Status"] = "Valid"
        row["ABC_XYZ"] = f"{row['Usage_Value_ABC']}{row['XYZ']}" if row["XYZ"] in {"X", "Y", "Z"} else "Unclassified"
    for row in excluded:
        row.update(
            {
                "Rank": None,
                "Annual_Usage_Value_Share": None,
                "Cumulative_Annual_Usage_Value_Share": None,
                "Usage_Value_ABC": "Unclassified",
                "Usage_Value_ABC_Boundary_Distance": None,
                "Segmentation_Status": "No positive usage value" if row["Complete_History"] == "Yes" else "Incomplete or invalid history",
                "ABC_XYZ": "Unclassified",
            }
        )
    quality = [
        quality_row("Demand rows", len(rows), "INFO", "Valid numeric/date rows"),
        quality_row("Demand item-locations", len(groups), "INFO", "Segmentation grain"),
        quality_row("Incomplete demand histories", incomplete, "PASS" if not incomplete else "FAIL", "Missing periods are not converted to zero"),
    ]
    return ranked + excluded, quality


def opportunity_flags(categories: list[dict[str, Any]], suppliers: list[dict[str, Any]], args: argparse.Namespace) -> list[dict[str, Any]]:
    output = []
    for row in categories:
        category = row["Category_L1"]
        if row["Supplier_Count"] >= args.fragmented_supplier_count:
            output.append({"Scope_Type": "Category", "Scope": category, "Flag": "Supplier fragmentation review", "Evidence": f"{row['Supplier_Count']} suppliers", "Positive_Spend": row["Positive_Spend"], "Savings_Claim": "None - validate addressability"})
        if row["Top_1_Supplier_Share"] is not None and row["Top_1_Supplier_Share"] >= args.concentration_share:
            output.append({"Scope_Type": "Category", "Scope": category, "Flag": "Supplier concentration review", "Evidence": f"Top supplier share {row['Top_1_Supplier_Share']:.1%}", "Positive_Spend": row["Positive_Spend"], "Savings_Claim": "None - assess continuity and market structure"})
        if row["Kraljic_Quadrant"] == "Unscored" and row["Spend_ABC"] == "A":
            output.append({"Scope_Type": "Category", "Scope": category, "Flag": "Kraljic evidence gap", "Evidence": "A-spend category lacks complete impact/risk score", "Positive_Spend": row["Positive_Spend"], "Savings_Claim": "None"})
        if row["Contract_Status_Coverage"] and row["On_Contract_Share_of_Known"] is not None and row["On_Contract_Share_of_Known"] < 0.8:
            output.append({"Scope_Type": "Category", "Scope": category, "Flag": "Contract leakage review", "Evidence": f"On-contract share of known spend {row['On_Contract_Share_of_Known']:.1%}", "Positive_Spend": row["Positive_Spend"], "Savings_Claim": "None - validate eligibility and cause"})
    for row in suppliers:
        if row["Transaction_Count"] == 1 and row["Positive_Spend"] > 0:
            output.append({"Scope_Type": "Supplier", "Scope": row["Supplier_Name"], "Flag": "One-time supplier / tail review", "Evidence": "One transaction in analysis scope", "Positive_Spend": row["Positive_Spend"], "Savings_Claim": "None - check criticality and business need"})
    return sorted(output, key=lambda row: (-float(row["Positive_Spend"]), row["Flag"], row["Scope"]))


def main() -> None:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    spend_headers, raw_spend = read_table(args.input, args.spend_sheet if args.input.suffix.lower() != ".csv" else None)
    require_columns(spend_headers, SPEND_REQUIRED, "Spend input")
    spend_rows, counters = normalize_spend(raw_spend)
    quality = spend_quality(len(raw_spend), spend_rows, counters, args)
    suppliers = supplier_summary(spend_rows, args)
    categories = category_summary(spend_rows, args)
    months = monthly_summary(spend_rows)
    flags = opportunity_flags(categories, suppliers, args)

    demand_source = None
    raw_demand: list[dict[str, Any]] = []
    demand_counters: dict[str, int] = {}
    segmentation: list[dict[str, Any]] = []
    if args.demand_file:
        demand_source = args.demand_file
        demand_headers, raw_demand = read_table(args.demand_file, args.demand_sheet if args.demand_file.suffix.lower() != ".csv" else None)
    elif workbook_has_sheet(args.input, args.demand_sheet):
        demand_source = args.input
        demand_headers, raw_demand = read_table(args.input, args.demand_sheet)
    else:
        demand_headers = []
    if raw_demand:
        require_columns(demand_headers, DEMAND_REQUIRED, "Demand input")
        demand_rows, demand_counters = normalize_demand(raw_demand)
        segmentation, demand_quality = item_segmentation(demand_rows, args)
        quality.extend(demand_quality)
        demand_currencies = sorted({text(row.get("Base_Currency")) for row in demand_rows if text(row.get("Base_Currency"))})
        spend_currencies = sorted({text(row.get("Base_Currency")) for row in spend_rows if text(row.get("Base_Currency"))})
        quality.extend(
            [
                quality_row("Demand base currency", ", ".join(demand_currencies) or "Missing", "PASS" if len(demand_currencies) == 1 and demand_currencies == spend_currencies else "FAIL", "Demand unit cost must use the same base currency as spend"),
                quality_row("Invalid demand period rows", demand_counters.get("invalid_period", 0), "PASS" if not demand_counters.get("invalid_period") else "FAIL", "Excluded from segmentation"),
                quality_row("Invalid demand/cost rows", demand_counters.get("invalid_numeric", 0), "PASS" if not demand_counters.get("invalid_numeric") else "FAIL", "Excluded from segmentation"),
                quality_row("Negative demand rows", demand_counters.get("negative_demand", 0), "PASS" if not demand_counters.get("negative_demand") else "FAIL", "Returns require explicit treatment"),
                quality_row("Nonpositive unit-cost rows", demand_counters.get("nonpositive_cost", 0), "PASS" if not demand_counters.get("nonpositive_cost") else "FAIL", "Annual usage value requires positive approved cost"),
                quality_row("Duplicate demand keys", demand_counters.get("duplicate_keys", 0), "PASS" if not demand_counters.get("duplicate_keys") else "FAIL", "Expected one part-location-period row"),
            ]
        )

    gross = sum(row["Positive_Spend"] for row in spend_rows)
    credits = sum(row["Credit_Spend"] for row in spend_rows)
    net = gross + credits
    currencies = sorted({text(row.get("Base_Currency")) for row in spend_rows if text(row.get("Base_Currency"))})
    failed_checks = [row["Check"] for row in quality if row["Status"] == "FAIL"]
    summary = [
        {
            "Analysis_Status": "BLOCKED" if failed_checks else "READY_WITH_REVIEW",
            "Base_Currency": currencies[0] if len(currencies) == 1 else "MIXED_OR_MISSING",
            "Scope_Start": min((row["Transaction_Date"] for row in spend_rows), default=""),
            "Scope_End": max((row["Transaction_Date"] for row in spend_rows), default=""),
            "Net_Spend": net,
            "Gross_Positive_Spend": gross,
            "Credits": credits,
            "Line_Count": len(spend_rows),
            "Transaction_Count": len({(text(row.get("Source_System")), text(row.get("Transaction_ID"))) for row in spend_rows}),
            "Supplier_Count": len(suppliers),
            "Category_Count": len(categories),
            "Demand_Item_Location_Count": len(segmentation),
            "Failed_Quality_Checks": "; ".join(failed_checks),
        }
    ]

    outputs = {
        "spend_summary.csv": summary,
        "supplier_summary.csv": suppliers,
        "category_summary.csv": categories,
        "monthly_spend.csv": months,
        "opportunity_flags.csv": flags,
        "data_quality.csv": quality,
    }
    if segmentation:
        outputs["item_segmentation.csv"] = segmentation
    for filename, rows in outputs.items():
        write_csv(args.output_dir / filename, rows)

    metadata = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "input": str(args.input),
        "spend_sheet": args.spend_sheet,
        "demand_source": str(demand_source) if demand_source else None,
        "demand_sheet": args.demand_sheet if demand_source else None,
        "analysis_status": summary[0]["Analysis_Status"],
        "failed_quality_checks": failed_checks,
        "basis": {
            "reconciliation": "signed net Spend_Base",
            "pareto": "gross positive Spend_Base; crossing item remains in the lower class",
            "supplier_entity": "parent, else normalized entity, else raw supplier",
        },
        "thresholds": {
            "abc_a": args.abc_a,
            "abc_b": args.abc_b,
            "xyz_x_cv": args.xyz_x,
            "xyz_y_cv": args.xyz_y,
            "adi": args.adi_boundary,
            "cv2": args.cv2_boundary,
            "kraljic": args.kraljic_threshold,
            "fragmented_supplier_count": args.fragmented_supplier_count,
            "concentration_share": args.concentration_share,
        },
        "outputs": sorted(outputs),
        "notes": [
            "Opportunity flags are review prompts and contain no savings estimate.",
            "Kraljic scores are aggregated only from supplied 0-100 evidence fields.",
            "XYZ and ADI/CV2 require complete demand periods; missing periods are not treated as zero.",
        ],
    }
    (args.output_dir / "analysis_metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    print(f"Wrote {len(outputs) + 1} outputs to {args.output_dir}")
    print(f"Analysis status: {summary[0]['Analysis_Status']}")
    if failed_checks:
        print("Failed checks: " + "; ".join(failed_checks))


if __name__ == "__main__":
    main()
