"""Validate the synthetic workbook contract used by Demand Genie version 3."""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "Demand_Genie_Synthetic_Demand_History.xlsx"
REQUIRED_SHEETS = {
    "Demand_Data", "Product_Master", "DDMRP_Positions", "DDMRP_Recommendations", "Spend_Lines",
    "Supplier_Master", "Contracts", "Category_Taxonomy", "Category_Risk", "Spend_Control", "Demand_History",
}


def records(workbook, name: str) -> list[dict[str, object]]:
    values = workbook[name].iter_rows(values_only=True)
    headers = list(next(values))
    return [dict(zip(headers, row)) for row in values]


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    assert not missing, f"Missing sheets: {', '.join(missing)}"

    demand = records(workbook, "Demand_Data")
    positions = records(workbook, "DDMRP_Positions")
    recommendations = records(workbook, "DDMRP_Recommendations")
    products = records(workbook, "Product_Master")
    spend = records(workbook, "Spend_Lines")
    demand_history = records(workbook, "Demand_History")
    category_risk = records(workbook, "Category_Risk")
    control = records(workbook, "Spend_Control")[0]

    assert len(demand) == 2880
    assert len({row["SKU"] for row in demand}) == 80
    assert len({row["Month"] for row in demand}) == 36
    assert len(positions) == len(recommendations) == len(products) == 80
    assert all(float(row["Unit_Cost_EUR"]) > 0 for row in products)

    spend_keys = [(row["Source_System"], row["Transaction_ID"], row["Line_ID"]) for row in spend]
    assert len(spend_keys) == len(set(spend_keys))
    assert {row["Base_Currency"] for row in spend} == {"EUR"}
    assert all(row["Supplier_ID"] and row["Supplier_Normalized_ID"] and row["Supplier_Parent_ID"] for row in spend)
    assert all(row["Category_L1"] and row["Category_Code"] for row in spend)
    assert all(str(row["On_Contract"]) in {"Yes", "No"} for row in spend)
    assert any(float(row["Spend_Base"]) < 0 for row in spend)
    assert any(float(row["Spend_Base"]) > 0 for row in spend)
    analyzed_total = sum(float(row["Spend_Base"]) for row in spend)
    control_total = float(control["Control_Net_Spend"])
    assert math.isclose(analyzed_total, control_total, rel_tol=1e-9, abs_tol=0.01)

    demand_keys = [(row["Part_ID"], row["Location"], row["Period"]) for row in demand_history]
    assert len(demand_history) == 2880
    assert len(demand_keys) == len(set(demand_keys))
    assert {row["Base_Currency"] for row in demand_history} == {"EUR"}
    assert all(float(row["Unit_Cost_Base"]) > 0 for row in demand_history)

    quadrants = {row["Kraljic_Quadrant"] for row in category_risk}
    assert quadrants == {"Strategic", "Leverage", "Bottleneck", "Routine"}
    assert all(0 <= float(row["Business_Impact_Score"]) <= 100 for row in category_risk)
    assert all(0 <= float(row["Supply_Risk_Score"]) <= 100 for row in category_risk)

    print(json.dumps({
        "status": "PASS",
        "demand_rows": len(demand),
        "ddmrp_positions": len(positions),
        "spend_lines": len(spend),
        "supplier_rows": workbook["Supplier_Master"].max_row - 1,
        "contract_rows": workbook["Contracts"].max_row - 1,
        "net_spend_eur": round(analyzed_total, 2),
        "reconciliation_difference": round(analyzed_total - control_total, 2),
        "kraljic_quadrants": sorted(quadrants),
    }, indent=2))


if __name__ == "__main__":
    main()
