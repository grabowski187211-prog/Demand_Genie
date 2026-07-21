"""Numerical and responsive browser tests for Demand Genie Dashboard V4."""

from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from playwright.sync_api import Page, sync_playwright


ROOT = Path(__file__).resolve().parents[1]
URL = sys.argv[1] if len(sys.argv) > 1 else "http://127.0.0.1:8001/dashboard-v4.html"
WORKBOOK = ROOT / "data" / "Demand_Genie_Synthetic_Demand_History.xlsx"
FORECAST_DIR = ROOT / "data" / "forecast-v4"


def make_adversarial_workbook(directory: Path) -> Path:
    path = directory / WORKBOOK.name
    shutil.copy2(WORKBOOK, path)
    workbook = load_workbook(path)
    demand = workbook["Demand_Data"]
    demand.cell(row=2, column=6).value = 99999
    positions = workbook["DDMRP_Positions"]
    positions.cell(row=2, column=5).value = None
    workbook.save(path)
    return path


def expected_for_first_sku() -> dict[str, object]:
    scores = pd.read_csv(FORECAST_DIR / "model_selection.csv")
    first_sku = sorted(scores["SKU"].unique())[0]
    selected = scores[(scores["SKU"] == first_sku) & scores["Selected"]].iloc[0]
    forecast = pd.read_csv(FORECAST_DIR / "forecast_results.csv")
    first_forecast = forecast[forecast["SKU"] == first_sku].sort_values("Month").iloc[0]
    return {
        "sku": first_sku,
        "model": selected["Selected_Model"],
        "rmse": f"{selected['RMSE']:,.1f}",
        "first_forecast": round(float(first_forecast["Forecast_Units"])),
    }


def verify_core(page: Page, viewport_name: str, expected: dict[str, object]) -> dict[str, object]:
    page.goto(URL, wait_until="load")
    page.wait_for_selector("#planning-body tr")
    assert page.locator(".tab").count() == 7
    assert page.locator("#planning-body tr").count() == 80
    assert page.locator("#source-file").inner_text() == WORKBOOK.name
    assert page.locator("#quality-gate-body tr").count() >= 9
    assert page.locator("#quality-gate-body").get_by_text("Forecast artifact provenance").count() == 1
    assert page.locator("#quality-gate-body").get_by_text("PASS", exact=True).count() >= 4

    first_tab = page.get_by_role("tab", name="Replenishment")
    first_tab.focus()
    page.keyboard.press("ArrowRight")
    assert page.get_by_role("tab", name="Forecast").get_attribute("aria-selected") == "true"

    page.wait_for_selector("#model-comparison .model-row")
    assert page.locator("#forecast-sku").input_value() == expected["sku"]
    assert page.locator("#selected-model").inner_text() == str(expected["model"]).replace("Seasonal_Naive", "Seasonal naive")
    assert page.locator("#forecast-rmse").inner_text() == expected["rmse"]
    assert page.locator("#model-comparison .model-row").count() == 7
    assert "42 errors" in page.locator("#model-context").inner_text()
    assert page.locator("#model-comparison").get_by_text("TiRex2", exact=False).count() == 1

    page.select_option("#forecast-model", "TiRex2")
    assert page.locator("#selected-model").inner_text() == "TiRex2"
    assert page.locator("#forecast-interval-legend").inner_text() == "native 80%"
    assert page.locator("#adjustment-interval-label").inner_text().lower() == "native 80% interval"
    assert "no native 95%" in page.locator("#forecast-review-note").inner_text().lower()

    page.locator('#horizon-control button[data-horizon="12"]').click()
    assert page.locator("#forecast-chart .svg-point.forecast").count() == 12
    with page.expect_download() as download_info:
        page.locator("#export-forecast").click()
    assert download_info.value.suggested_filename == f"demand-genie-v4-forecast-{expected['sku']}.csv"

    page.locator('#forecast-mode-control button[data-forecast-mode="decomposition"]').click()
    assert page.locator("#decomposition-results").is_visible()
    assert page.locator("#decomposition-chart .decomp-line").count() == 4
    assert page.locator("#seasonal-cycles").inner_text() == "3.0"
    assert page.locator("#trend-strength").inner_text() != "-"

    page.get_by_role("tab", name="Spend").click()
    page.wait_for_selector("#pareto-chart .pareto-bar")
    assert page.locator("#pareto-chart .pareto-bar").count() == 10
    assert page.locator("#pareto-chart .pareto-threshold-label").text_content() == "80% cumulative"
    initial_spend = page.locator("#net-spend").inner_text()
    page.select_option("#spend-period-filter", "Current")
    assert page.locator("#net-spend").inner_text() != initial_spend
    with page.expect_download() as spend_download:
        page.locator("#export-spend").click()
    assert spend_download.value.suggested_filename == "demand-genie-v4-spend-analysis.csv"

    page.get_by_role("tab", name="Portfolio").click()
    page.wait_for_selector("#kraljic-chart .portfolio-point")
    assert page.locator("#kraljic-chart .portfolio-point").count() == 10
    chart_width = page.locator("#kraljic-chart").evaluate("node => node.getBoundingClientRect().width")
    container_width = page.locator(".kraljic-wrap").evaluate("node => node.getBoundingClientRect().width")
    assert chart_width <= container_width + 1

    page.get_by_role("tab", name="Execution").click()
    assert page.locator("#open-supply-total").inner_text() == "99"

    page.get_by_role("tab", name="Data").click()
    assert page.locator("#spend-line-count").inner_text() == "1,964"
    assert "warning" in page.locator("#data-headline").inner_text().lower()

    overflow = page.evaluate("document.documentElement.scrollWidth - document.documentElement.clientWidth")
    assert overflow <= 2, f"{viewport_name} horizontal overflow: {overflow}px"
    return {
        "viewport": viewport_name,
        "model": expected["model"],
        "model_rows": page.locator("#model-comparison .model-row").count(),
        "horizontal_overflow_px": overflow,
    }


def verify_upload_hash_gate(page: Page, adversarial: Path) -> None:
    page.goto(URL, wait_until="load")
    page.set_input_files("#workbook-input", str(WORKBOOK))
    page.wait_for_function("document.querySelector('#source-kind').textContent.trim() === 'Uploaded workbook'")
    page.get_by_role("tab", name="Forecast").click()
    assert not page.locator('#forecast-model option[value="TiRex2"]').is_disabled()
    assert page.locator("#model-comparison .model-row").count() == 7

    page.set_input_files("#workbook-input", str(adversarial))
    page.wait_for_function("document.querySelector('#data-alert').textContent.includes('No content-matched offline forecast package')")
    assert page.locator('#forecast-model option[value="TiRex2"]').is_disabled()
    assert page.locator("#model-comparison .model-row").count() == 5
    page.get_by_role("tab", name="Replenishment").click()
    assert page.locator("#planning-body .status-data").count() == 1
    assert "blocked" in page.locator("#planning-headline").inner_text().lower()


def main() -> None:
    errors: list[str] = []
    results = []
    expected = expected_for_first_sku()
    screenshots: list[str] = []
    with tempfile.TemporaryDirectory() as temp_dir:
        adversarial = make_adversarial_workbook(Path(temp_dir))
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch()
            for name, viewport in (
                ("desktop", {"width": 1440, "height": 1000}),
                ("mobile", {"width": 390, "height": 844}),
            ):
                page = browser.new_page(viewport=viewport)
                page.on("console", lambda message: errors.append(f"console {message.type}: {message.text}") if message.type == "error" else None)
                page.on("pageerror", lambda error: errors.append(f"pageerror: {error}"))
                results.append(verify_core(page, name, expected))
                page.get_by_role("tab", name="Forecast").click()
                page.locator('#forecast-mode-control button[data-forecast-mode="decomposition"]').click()
                screenshot = f"/tmp/demand-genie-v4-{name}-decomposition.png"
                page.screenshot(path=screenshot, full_page=True)
                screenshots.append(screenshot)
                page.get_by_role("tab", name="Data").click()
                screenshot = f"/tmp/demand-genie-v4-{name}-data.png"
                page.screenshot(path=screenshot, full_page=True)
                screenshots.append(screenshot)
                page.close()

            upload_page = browser.new_page(viewport={"width": 1200, "height": 900})
            upload_page.on("console", lambda message: errors.append(f"console {message.type}: {message.text}") if message.type == "error" else None)
            upload_page.on("pageerror", lambda error: errors.append(f"pageerror: {error}"))
            verify_upload_hash_gate(upload_page, adversarial)
            upload_page.close()
            browser.close()

    assert not errors, "\n".join(errors)
    print(json.dumps({"status": "PASS", "results": results, "screenshots": screenshots}, indent=2))


if __name__ == "__main__":
    main()
