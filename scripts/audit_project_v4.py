#!/usr/bin/env python3
"""Independently audit Demand Genie V4 source data, calculations, and artifacts."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import statistics
import sys
from collections import defaultdict
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data" / "Demand_Genie_Synthetic_Demand_History.xlsx"
FORECAST_DIR = ROOT / "data" / "forecast-v4"
SPEND_DIR = ROOT / "data" / "spend-analysis"
OUTPUT_PATH = ROOT / "data" / "audit-v4.json"

EXPECTED_DASHBOARD_HASHES = {
    "index.html": "ece5a062b805e19083ccb3ebcd0cac912b25cdc2f54e734c81e3eab223b808a9",
    "dashboard-v2.html": "24d843bd2bd2c31634d95024ea0d72973af094e9d08d9f3c6af95def9fca5b4e",
    "dashboard-v3.html": "7875d71c9d7f5e740e5cfd8b99f9ba4107a6fd520f15102c56c1cbd1ea82d035",
}
MODEL_COMPLEXITY = {
    "Mean": 1,
    "Naive": 1,
    "Seasonal_Naive": 1,
    "Drift": 2,
    "ETS": 3,
    "ARIMA": 4,
    "TiRex2": 5,
}
NUMERIC_TOLERANCE = 1e-8
MONEY_TOLERANCE = 0.01


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def workbook_records(workbook, sheet_name: str) -> list[dict[str, Any]]:
    values = workbook[sheet_name].iter_rows(values_only=True)
    headers = [str(value) for value in next(values)]
    return [dict(zip(headers, row)) for row in values]


def csv_records(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as source:
        return list(csv.DictReader(source))


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def integer(value: Any) -> int:
    parsed = number(value)
    if parsed is None:
        raise ValueError(f"Expected integer-compatible value, received {value!r}")
    return int(round(parsed))


def month_key(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    return str(value)[:7]


def add_months(month: str, count: int) -> str:
    year, month_number = (int(part) for part in month.split("-"))
    index = year * 12 + month_number - 1 + count
    return f"{index // 12:04d}-{index % 12 + 1:02d}"


def is_close(actual: Any, expected: Any, *, abs_tol: float = NUMERIC_TOLERANCE) -> bool:
    left = number(actual)
    right = number(expected)
    if left is None or right is None:
        return left is None and right is None
    return math.isclose(left, right, rel_tol=1e-9, abs_tol=abs_tol)


def bool_value(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "1"}:
        return True
    if normalized in {"false", "no", "0"}:
        return False
    return None


class Audit:
    def __init__(self) -> None:
        self.checks: list[dict[str, Any]] = []
        self.warnings: list[dict[str, Any]] = []
        self.metrics: dict[str, Any] = {}

    def check(self, name: str, passed: bool, detail: str) -> None:
        self.checks.append({"name": name, "status": "PASS" if passed else "FAIL", "detail": detail})

    def warn(self, name: str, detail: str, count: int | float | None = None) -> None:
        row: dict[str, Any] = {"name": name, "status": "WARN", "detail": detail}
        if count is not None:
            row["count"] = count
        self.warnings.append(row)

    @property
    def failures(self) -> list[dict[str, Any]]:
        return [row for row in self.checks if row["status"] == "FAIL"]


def compare_fields(
    actual: dict[str, Any], expected: dict[str, Any], fields: Iterable[str], *, abs_tol: float = NUMERIC_TOLERANCE
) -> list[str]:
    return [field for field in fields if not is_close(actual.get(field), expected.get(field), abs_tol=abs_tol)]


def pareto_rows(rows: list[dict[str, Any]], value_field: str) -> list[dict[str, Any]]:
    ordered = sorted(rows, key=lambda row: (-max(0.0, float(row[value_field])), str(row)))
    total = sum(max(0.0, float(row[value_field])) for row in ordered)
    cumulative = 0.0
    for rank, row in enumerate(ordered, 1):
        value = max(0.0, float(row[value_field]))
        prior_share = cumulative / total if total else 0.0
        cumulative += value
        row["Rank"] = rank
        row["Share"] = value / total if total else None
        row["Cumulative"] = cumulative / total if total else None
        row["ABC"] = "A" if prior_share < 0.8 else "B" if prior_share < 0.95 else "C"
    return ordered


def audit_immutable_dashboards(audit: Audit) -> None:
    mismatches = []
    for filename, expected_hash in EXPECTED_DASHBOARD_HASHES.items():
        actual_hash = sha256_file(ROOT / filename)
        if actual_hash != expected_hash:
            mismatches.append(filename)
    audit.check(
        "Versions 1-3 are byte-for-byte unchanged",
        not mismatches,
        "Immutable hashes match" if not mismatches else f"Hash mismatch: {', '.join(mismatches)}",
    )


def audit_demand(audit: Audit, sheets: dict[str, list[dict[str, Any]]]) -> dict[str, list[tuple[str, float]]]:
    demand = sheets["Demand_Data"]
    grouped: defaultdict[str, list[tuple[str, float]]] = defaultdict(list)
    keys = []
    for row in demand:
        key = (str(row["SKU"]), month_key(row["Month"]))
        keys.append(key)
        value = number(row["Demand_Units"])
        if value is not None:
            grouped[key[0]].append((key[1], value))

    structural_errors = []
    for sku, history in grouped.items():
        history.sort()
        months = [month for month, _ in history]
        expected = [add_months(months[0], index) for index in range(len(months))] if months else []
        if len(history) != 36 or months != expected:
            structural_errors.append(sku)
    nonnegative = all(value >= 0 for history in grouped.values() for _, value in history)
    audit.check(
        "Demand history has complete unique monthly keys",
        len(demand) == 2880 and len(grouped) == 80 and len(keys) == len(set(keys)) and not structural_errors,
        f"{len(demand):,} rows; {len(grouped)} SKUs; 36 contiguous months per SKU",
    )
    audit.check("Demand values are numeric and nonnegative", nonnegative, "All demand values are usable")

    demand_history = sheets["Demand_History"]
    history_map = {
        (str(row["Part_ID"]), month_key(row["Period"])): number(row["Demand_Units"])
        for row in demand_history
    }
    aligned = len(history_map) == len(demand) and all(
        is_close(history_map.get((str(row["SKU"]), month_key(row["Month"]))), row["Demand_Units"])
        for row in demand
    )
    audit.check("Forecast and spend demand facts reconcile", aligned, "Demand_Data equals Demand_History at SKU-month grain")

    products = {str(row["SKU"]): row for row in sheets["Product_Master"]}
    intermittent_without_zeros = sum(
        1
        for sku, history in grouped.items()
        if str(products[sku]["Demand_Profile"]) == "Intermittent" and all(value > 0 for _, value in history)
    )
    if intermittent_without_zeros:
        audit.warn(
            "Synthetic intermittency is amplitude-only",
            "The profile label says intermittent, but every monthly observation is positive; do not treat it as sparse demand.",
            intermittent_without_zeros,
        )
    audit.metrics.update({"demand_rows": len(demand), "sku_count": len(grouped), "history_months": 36})
    return grouped


def buffer_status(net_flow: float, top_red: float, top_yellow: float, top_green: float) -> str:
    if net_flow < 0:
        return "Black"
    if net_flow <= top_red:
        return "Red"
    if net_flow <= top_yellow:
        return "Yellow"
    if net_flow <= top_green:
        return "Green"
    return "Blue"


def round_up(value: float, multiple: int) -> int:
    return int(math.ceil(value / multiple) * multiple) if value > 0 else 0


def audit_ddmrp(
    audit: Audit,
    sheets: dict[str, list[dict[str, Any]]],
    histories: dict[str, list[tuple[str, float]]],
) -> None:
    positions = {(str(row["SKU"]), str(row["Location"])): row for row in sheets["DDMRP_Positions"]}
    recommendations = {(str(row["SKU"]), str(row["Location"])): row for row in sheets["DDMRP_Recommendations"]}
    master = {(str(row["SKU"]), str(row["Location"])): row for row in sheets["DDMRP_Master"]}

    inventory: defaultdict[tuple[str, str], float] = defaultdict(float)
    for row in sheets["Inventory_Snapshot"]:
        inventory[(str(row["SKU"]), str(row["Location"]))] += number(row["Usable_On_Hand_Units"]) or 0.0
    supply: defaultdict[tuple[str, str], float] = defaultdict(float)
    late_eligible = 0
    for row in sheets["Open_Supply"]:
        if bool_value(row["Eligible_for_NFP"]):
            supply[(str(row["SKU"]), str(row["Location"]))] += number(row["Open_Supply_Units"]) or 0.0
            if str(row["Order_Status"]) == "Late":
                late_eligible += 1
    qualified: defaultdict[tuple[str, str], defaultdict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in sheets["Qualified_Demand"]:
        if bool_value(row["Qualified_for_NFP"]):
            qualified[(str(row["SKU"]), str(row["Location"]))][str(row["Demand_Type"])] += number(row["Demand_Units"]) or 0.0

    source_mismatches: list[str] = []
    formula_mismatches: list[str] = []
    adu_mismatches: list[str] = []
    status_counts: defaultdict[str, int] = defaultdict(int)
    required_position_fields = {
        "ADU", "DLT_Days", "Lead_Time_Factor", "Variability_Factor", "MOQ", "Order_Cycle_Days",
        "On_Hand", "Open_Supply", "Past_Due_Demand", "Today_Demand", "Qualified_Spike_Demand",
        "Order_Multiple", "Demand_Adjustment_Factor",
    }
    missing_or_invalid = 0

    for key, position in positions.items():
        if any(number(position.get(field)) is None for field in required_position_fields):
            missing_or_invalid += 1
            continue
        sku, _ = key
        history = sorted(histories[sku])
        trailing = history[-12:]
        days = sum(
            (date(int(add_months(month, 1)[:4]), int(add_months(month, 1)[5:]), 1) - date(int(month[:4]), int(month[5:]), 1)).days
            for month, _ in trailing
        )
        expected_adu = sum(value for _, value in trailing) / days
        if not is_close(position["ADU"], expected_adu) or not is_close(master[key]["Baseline_ADU"], expected_adu):
            adu_mismatches.append(sku)

        expected_sources = {
            "On_Hand": inventory[key],
            "Open_Supply": supply[key],
            "Past_Due_Demand": qualified[key]["Past_Due"],
            "Today_Demand": qualified[key]["Today"],
            "Qualified_Spike_Demand": qualified[key]["Qualified_Spike"],
        }
        if compare_fields(position, expected_sources, expected_sources):
            source_mismatches.append(sku)

        adu = float(position["ADU"])
        adjustment = float(position["Demand_Adjustment_Factor"])
        dlt = float(position["DLT_Days"])
        ltf = float(position["Lead_Time_Factor"])
        vf = float(position["Variability_Factor"])
        moq = float(position["MOQ"])
        cycle = float(position["Order_Cycle_Days"])
        multiple = integer(position["Order_Multiple"])
        adjusted_adu = adu * adjustment
        yellow = adjusted_adu * dlt
        red_base = yellow * ltf
        red_safety = red_base * vf
        red = red_base + red_safety
        green = max(moq, adjusted_adu * cycle, adjusted_adu * dlt * ltf)
        top_yellow = red + yellow
        top_green = top_yellow + green
        qualified_total = (
            float(position["Past_Due_Demand"])
            + float(position["Today_Demand"])
            + float(position["Qualified_Spike_Demand"])
        )
        net_flow = float(position["On_Hand"]) + float(position["Open_Supply"]) - qualified_total
        triggered = net_flow <= top_yellow
        base_order = max(0.0, top_green - net_flow) if triggered else 0.0
        expected = {
            "Adjusted_ADU": adjusted_adu,
            "Red_Base": red_base,
            "Red_Safety": red_safety,
            "Red_Zone": red,
            "Yellow_Zone": yellow,
            "Green_Zone": green,
            "TOR": red,
            "TOY": top_yellow,
            "TOG": top_green,
            "Qualified_Demand": qualified_total,
            "Net_Flow_Position": net_flow,
            "NFP_Percent_of_TOG": net_flow / top_green * 100.0,
            "Base_Recommended_Units": base_order,
            "Recommended_Order_Units": round_up(base_order, multiple),
            "On_Hand_Gap_to_TOR": float(position["On_Hand"]) - red,
        }
        recommendation = recommendations[key]
        mismatched_fields = compare_fields(recommendation, expected, expected)
        expected_status = buffer_status(net_flow, red, top_yellow, top_green)
        status_counts[expected_status] += 1
        if str(recommendation["NFP_Status"]) != expected_status:
            mismatched_fields.append("NFP_Status")
        if bool_value(recommendation["Replenishment_Triggered"]) != triggered:
            mismatched_fields.append("Replenishment_Triggered")
        if mismatched_fields:
            formula_mismatches.append(f"{sku}: {', '.join(mismatched_fields)}")

    audit.check(
        "DDMRP position inputs are complete",
        len(positions) == 80 and len(recommendations) == 80 and not missing_or_invalid,
        f"{len(positions)} item-location positions; {missing_or_invalid} invalid input rows",
    )
    audit.check(
        "DDMRP source facts aggregate to positions",
        not source_mismatches,
        "Inventory, eligible open supply, and qualified demand reconcile" if not source_mismatches else f"Mismatches: {source_mismatches[:5]}",
    )
    audit.check(
        "DDMRP ADU uses trailing 12 complete months/calendar days",
        not adu_mismatches,
        "All 80 baseline ADU values independently recomputed" if not adu_mismatches else f"Mismatches: {adu_mismatches[:5]}",
    )
    audit.check(
        "DDMRP zones, NFP, trigger, status, and order quantities reconcile",
        not formula_mismatches,
        "All recommendations match independent formulas" if not formula_mismatches else formula_mismatches[0],
    )
    if late_eligible:
        audit.warn(
            "Late supply remains eligible for NFP",
            "This is an explicit planning policy requiring execution review, not a calculation error.",
            late_eligible,
        )
    audit.metrics["ddmrp"] = {
        "positions": len(positions),
        "status_counts": dict(sorted(status_counts.items())),
        "late_eligible_supply_lines": late_eligible,
    }


def spend_parent(row: dict[str, Any]) -> str:
    return str(row.get("Supplier_Parent_ID") or row.get("Supplier_Normalized_ID") or row.get("Supplier_ID") or "UNMAPPED")


def audit_pareto_summary(
    source_rows: list[dict[str, Any]], output_rows: list[dict[str, str]], group_field: str, output_key: str
) -> list[str]:
    grouped: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in source_rows:
        grouped[str(row[group_field])].append(row)
    calculated = []
    for key, rows in grouped.items():
        calculated.append(
            {
                output_key: key,
                "Net_Spend": sum(float(row["Spend_Base"]) for row in rows),
                "Positive_Spend": sum(max(0.0, float(row["Spend_Base"])) for row in rows),
                "Credits": sum(min(0.0, float(row["Spend_Base"])) for row in rows),
            }
        )
    calculated = pareto_rows(calculated, "Positive_Spend")
    output_map = {str(row[output_key]): row for row in output_rows}
    mismatches = []
    for expected in calculated:
        actual = output_map.get(str(expected[output_key]))
        if actual is None:
            mismatches.append(str(expected[output_key]))
            continue
        field_pairs = {
            "Net_Spend": "Net_Spend",
            "Positive_Spend": "Positive_Spend",
            "Credits": "Credits",
            "Rank": "Rank",
            "Share": "Positive_Spend_Share",
            "Cumulative": "Cumulative_Positive_Spend_Share",
        }
        if any(not is_close(actual[target], expected[source], abs_tol=MONEY_TOLERANCE if "Spend" in source else NUMERIC_TOLERANCE) for source, target in field_pairs.items()):
            mismatches.append(str(expected[output_key]))
        if str(actual["Spend_ABC"]) != expected["ABC"]:
            mismatches.append(str(expected[output_key]))
    return sorted(set(mismatches))


def audit_spend(audit: Audit, sheets: dict[str, list[dict[str, Any]]]) -> None:
    spend = sheets["Spend_Lines"]
    control = sheets["Spend_Control"][0]
    net = sum(float(row["Spend_Base"]) for row in spend)
    gross = sum(max(0.0, float(row["Spend_Base"])) for row in spend)
    credits = sum(min(0.0, float(row["Spend_Base"])) for row in spend)
    control_total = float(control["Control_Net_Spend"])
    currencies = {str(row["Base_Currency"]) for row in spend}
    keys = [(str(row["Source_System"]), str(row["Transaction_ID"]), str(row["Line_ID"])) for row in spend]
    reconciliation_ok = (
        abs(net - control_total) <= MONEY_TOLERANCE
        and abs(net - float(control["Analyzed_Net_Spend"])) <= MONEY_TOLERANCE
        and abs(float(control["Difference"])) <= MONEY_TOLERANCE
        and abs(gross - float(control["Gross_Positive_Spend"])) <= MONEY_TOLERANCE
        and abs(credits - float(control["Credits"])) <= MONEY_TOLERANCE
    )
    audit.check(
        "Signed spend ledger passes strict EUR 0.01 reconciliation",
        reconciliation_ok and currencies == {"EUR"},
        f"Net EUR {net:,.2f}; gross EUR {gross:,.2f}; credits EUR {credits:,.2f}; difference EUR {net - control_total:,.2f}",
    )
    audit.check(
        "Spend transaction-line keys are unique and complete",
        len(keys) == len(set(keys)) and all(all(key) for key in keys),
        f"{len(keys):,} signed ledger lines",
    )

    normalized = []
    for row in spend:
        normalized_row = dict(row)
        normalized_row["Supplier_Analysis_ID"] = spend_parent(row)
        normalized.append(normalized_row)
    category_source = [dict(row, Category_Group=str(row["Category_L1"])) for row in normalized]
    supplier_source = [dict(row, Supplier_Group=str(row["Supplier_Analysis_ID"])) for row in normalized]
    category_mismatches = audit_pareto_summary(
        category_source, csv_records(SPEND_DIR / "category_summary.csv"), "Category_Group", "Category_L1"
    )
    supplier_mismatches = audit_pareto_summary(
        supplier_source, csv_records(SPEND_DIR / "supplier_summary.csv"), "Supplier_Group", "Supplier_ID"
    )
    audit.check(
        "Category and supplier Pareto calculations reconcile",
        not category_mismatches and not supplier_mismatches,
        "Positive-spend shares, cumulative ranks, and ABC classes match" if not category_mismatches and not supplier_mismatches else f"Category {category_mismatches[:3]}; supplier {supplier_mismatches[:3]}",
    )

    months = sorted({month_key(row["Transaction_Date"]) for row in spend})
    expected_months = [add_months(months[0], index) for index in range(len(months))]
    monthly_output = {row["Month"]: row for row in csv_records(SPEND_DIR / "monthly_spend.csv")}
    monthly_ok = months == expected_months and len(monthly_output) == len(months)
    for month in months:
        month_rows = [row for row in spend if month_key(row["Transaction_Date"]) == month]
        expected = {
            "Net_Spend": sum(float(row["Spend_Base"]) for row in month_rows),
            "Positive_Spend": sum(max(0.0, float(row["Spend_Base"])) for row in month_rows),
            "Credits": sum(min(0.0, float(row["Spend_Base"])) for row in month_rows),
        }
        if month not in monthly_output or compare_fields(monthly_output[month], expected, expected, abs_tol=MONEY_TOLERANCE):
            monthly_ok = False
    audit.check("Monthly spend series is complete and reconciled", monthly_ok, f"{len(months)} contiguous ledger months")

    category_output = csv_records(SPEND_DIR / "category_summary.csv")
    category_risk = {str(row["Category_L1"]): row for row in sheets["Category_Risk"]}
    kraljic_errors = []
    near_threshold = 0
    for row in category_output:
        impact = float(row["Business_Impact_Score"])
        risk = float(row["Supply_Risk_Score"])
        expected = "Strategic" if impact >= 50 and risk >= 50 else "Leverage" if impact >= 50 else "Bottleneck" if risk >= 50 else "Routine"
        if str(row["Kraljic_Quadrant"]) != expected:
            kraljic_errors.append(str(row["Category_L1"]))
        if min(abs(impact - 50), abs(risk - 50)) <= 5:
            near_threshold += 1
        source = category_risk[str(row["Category_L1"])]
        if not is_close(impact, source["Business_Impact_Score"]) or not is_close(risk, source["Supply_Risk_Score"]):
            kraljic_errors.append(str(row["Category_L1"]))
    audit.check(
        "Kraljic scores and quadrants reconcile to factor evidence",
        not kraljic_errors,
        "All 10 category scores and 50/50 quadrant assignments match",
    )
    if near_threshold:
        audit.warn(
            "Kraljic threshold sensitivity",
            "At least one category is within five score points of a quadrant boundary; treat classification as a discussion aid.",
            near_threshold,
        )

    commitments = sum(float(row["Committed_Annual_Value_EUR"]) for row in sheets["Contracts"] if str(row["Contract_Status"]) == "Active")
    current_months = set(months[-12:])
    current_on_contract = sum(
        max(0.0, float(row["Spend_Base"]))
        for row in spend
        if month_key(row["Transaction_Date"]) in current_months and str(row["On_Contract"]).strip().lower() == "yes"
    )
    commitment_ratio = current_on_contract / commitments if commitments else math.inf
    if commitment_ratio > 1.5:
        audit.warn(
            "Synthetic contract commitments are implausibly low",
            f"Current 12-month positive on-contract spend is {commitment_ratio:.2f}x active annual commitments.",
            round(commitment_ratio, 2),
        )
    audit.warn(
        "Synthetic control is not independent",
        "Spend_Control is generated from the same ledger; it proves arithmetic reconciliation, not source-system completeness.",
    )
    audit.metrics["spend"] = {
        "line_count": len(spend),
        "net_eur": round(net, 2),
        "gross_positive_eur": round(gross, 2),
        "credits_eur": round(credits, 2),
        "contract_commitment_ratio": round(commitment_ratio, 4),
    }


def audit_item_segmentation(audit: Audit, sheets: dict[str, list[dict[str, Any]]]) -> None:
    rows = sheets["Demand_History"]
    grouped: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    periods = sorted({month_key(row["Period"]) for row in rows})
    for row in rows:
        grouped[(str(row["Part_ID"]), str(row["Location"]))].append(row)
    calculated = []
    for (part, location), members in grouped.items():
        ordered = sorted(members, key=lambda row: month_key(row["Period"]))
        values = [float(row["Demand_Units"]) for row in ordered]
        nonzero = [value for value in values if value > 0]
        mean = statistics.fmean(values)
        cv = statistics.pstdev(values) / mean if mean else None
        adi = len(values) / len(nonzero) if nonzero else None
        nonzero_mean = statistics.fmean(nonzero) if nonzero else None
        cv2 = (statistics.pstdev(nonzero) / nonzero_mean) ** 2 if nonzero and nonzero_mean else None
        xyz = "X" if cv is not None and cv < 0.5 else "Y" if cv is not None and cv < 1.0 else "Z" if cv is not None else "No demand"
        pattern = (
            "Smooth" if adi is not None and cv2 is not None and adi < 1.32 and cv2 < 0.49
            else "Erratic" if adi is not None and adi < 1.32
            else "Intermittent" if cv2 is not None and cv2 < 0.49
            else "Lumpy"
        )
        annual_units = mean * 12
        annual_value = annual_units * float(ordered[-1]["Unit_Cost_Base"])
        calculated.append(
            {
                "Part_ID": part,
                "Location": location,
                "Period_Count": len(values),
                "Mean_Demand": mean,
                "Zero_Share": (len(values) - len(nonzero)) / len(values),
                "Demand_CV": cv,
                "XYZ": xyz,
                "ADI": adi,
                "CV2_Nonzero_Demand": cv2,
                "Demand_Pattern": pattern,
                "Annualized_Usage_Units": annual_units,
                "Annual_Usage_Value": annual_value,
            }
        )
    calculated = pareto_rows(calculated, "Annual_Usage_Value")
    output = {(row["Part_ID"], row["Location"]): row for row in csv_records(SPEND_DIR / "item_segmentation.csv")}
    mismatches = []
    legacy_xyz = {str(row["SKU"]): str(row["XYZ_Class"]) for row in sheets["Product_Master"]}
    legacy_mismatches = 0
    for expected in calculated:
        key = (expected["Part_ID"], expected["Location"])
        actual = output.get(key)
        if actual is None:
            mismatches.append(expected["Part_ID"])
            continue
        numeric_fields = (
            "Period_Count", "Mean_Demand", "Zero_Share", "Demand_CV", "ADI", "CV2_Nonzero_Demand",
            "Annualized_Usage_Units", "Annual_Usage_Value",
        )
        if compare_fields(actual, expected, numeric_fields) or str(actual["XYZ"]) != expected["XYZ"] or str(actual["Demand_Pattern"]) != expected["Demand_Pattern"]:
            mismatches.append(expected["Part_ID"])
        if not is_close(actual["Rank"], expected["Rank"]) or not is_close(actual["Annual_Usage_Value_Share"], expected["Share"]) or not is_close(actual["Cumulative_Annual_Usage_Value_Share"], expected["Cumulative"]):
            mismatches.append(expected["Part_ID"])
        if str(actual["Usage_Value_ABC"]) != expected["ABC"]:
            mismatches.append(expected["Part_ID"])
        if legacy_xyz.get(expected["Part_ID"]) != expected["XYZ"]:
            legacy_mismatches += 1
    audit.check(
        "ABC/XYZ and ADI/CV2 segmentation independently reconcile",
        not mismatches and len(calculated) == 80 and len(periods) == 36,
        "All 80 item-location calculations match" if not mismatches else f"Mismatches: {sorted(set(mismatches))[:5]}",
    )
    if legacy_mismatches:
        audit.warn(
            "Legacy Product_Master XYZ labels conflict with measured variability",
            "V4 uses calculated 36-month segmentation and treats the master labels as legacy metadata.",
            legacy_mismatches,
        )


def forecast_metrics(rows: list[dict[str, str]]) -> dict[str, float | int | None]:
    errors = [float(row["Error"]) for row in rows]
    absolute = [abs(value) for value in errors]
    squared = [value * value for value in errors]
    actual = [float(row["Actual_Units"]) for row in rows]
    covered_80 = [bool_value(row["Covered_80"]) for row in rows]
    covered_95 = [bool_value(row["Covered_95"]) for row in rows]
    return {
        "ME": statistics.fmean(errors),
        "RMSE": math.sqrt(statistics.fmean(squared)),
        "MAE": statistics.fmean(absolute),
        "MASE": statistics.fmean(float(row["Scaled_Absolute_Error"]) for row in rows),
        "RMSSE": math.sqrt(statistics.fmean(float(row["Scaled_Squared_Error"]) for row in rows)),
        "WAPE": sum(absolute) / sum(actual),
        "Coverage_80": sum(value is True for value in covered_80) / len(rows),
        "Coverage_95": None if all(value is None for value in covered_95) else sum(value is True for value in covered_95) / len(rows),
        "Error_Sum": sum(errors),
        "Absolute_Error_Sum": sum(absolute),
        "Squared_Error_Sum": sum(squared),
        "Actual_Units_Sum": sum(actual),
        "Covered_80_Count": sum(value is True for value in covered_80),
        "Covered_95_Count": sum(value is True for value in covered_95),
    }


def audit_forecasts(
    audit: Audit,
    histories: dict[str, list[tuple[str, float]]],
) -> None:
    rolling = csv_records(FORECAST_DIR / "rolling_origin_predictions.csv")
    scores = csv_records(FORECAST_DIR / "model_selection.csv")
    all_forecasts = csv_records(FORECAST_DIR / "all_model_forecasts.csv")
    selected_forecasts = csv_records(FORECAST_DIR / "forecast_results.csv")
    grouped: defaultdict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rolling:
        grouped[(row["SKU"], row["Model"])].append(row)

    protocol_errors: list[str] = []
    score_errors: list[str] = []
    interval_errors: list[str] = []
    models = sorted({model for _, model in grouped})
    for (sku, model), rows in grouped.items():
        rows.sort(key=lambda row: (integer(row["Origin"]), integer(row["Horizon_Step"])))
        history = sorted(histories[sku])
        if len(rows) != 42 or {integer(row["Origin"]) for row in rows} != set(range(1, 8)):
            protocol_errors.append(f"{sku}/{model}: row/origin count")
            continue
        for row in rows:
            origin = integer(row["Origin"])
            training_months = integer(row["Training_Months"])
            step = integer(row["Horizon_Step"])
            expected_training = 23 + origin
            target_index = training_months + step - 1
            target_month, target_actual = history[target_index]
            train_values = [value for _, value in history[:training_months]]
            seasonal_differences = [train_values[index] - train_values[index - 12] for index in range(12, len(train_values))]
            mase_scale = statistics.fmean(abs(value) for value in seasonal_differences)
            rmsse_scale = statistics.fmean(value * value for value in seasonal_differences)
            forecast = float(row["Forecast_Units"])
            error = target_actual - forecast
            if (
                training_months != expected_training
                or step not in range(1, 7)
                or month_key(row["Month"]) != target_month
                or not is_close(row["Actual_Units"], target_actual)
                or not is_close(row["MASE_Scale"], mase_scale)
                or not is_close(row["RMSSE_Scale"], rmsse_scale)
                or not is_close(row["Error"], error)
                or not is_close(row["Absolute_Error"], abs(error))
                or not is_close(row["Squared_Error"], error * error)
                or not is_close(row["Scaled_Absolute_Error"], abs(error) / mase_scale)
                or not is_close(row["Scaled_Squared_Error"], error * error / rmsse_scale)
            ):
                protocol_errors.append(f"{sku}/{model}/origin {origin}/step {step}")
                break

            lower80, upper80 = number(row["PI80_Lower_Units"]), number(row["PI80_Upper_Units"])
            lower95, upper95 = number(row["PI95_Lower_Units"]), number(row["PI95_Upper_Units"])
            if lower80 is None or upper80 is None or lower80 > forecast or forecast > upper80:
                interval_errors.append(f"{sku}/{model}: 80%")
            if model == "TiRex2":
                quantiles = [number(row[f"Q{level}_Units"]) for level in range(10, 100, 10)]
                if any(value is None for value in quantiles) or any(left > right for left, right in zip(quantiles, quantiles[1:])):
                    interval_errors.append(f"{sku}/{model}: quantiles")
                if lower95 is not None or upper95 is not None or bool_value(row["Covered_95"]) is not None:
                    interval_errors.append(f"{sku}/{model}: false 95%")
            elif lower95 is None or upper95 is None or lower95 > lower80 or upper95 < upper80:
                interval_errors.append(f"{sku}/{model}: 95%")

        calculated = forecast_metrics(rows)
        actual_score = next((row for row in scores if row["SKU"] == sku and row["Model"] == model), None)
        if actual_score is None:
            score_errors.append(f"{sku}/{model}: missing")
        else:
            for field, expected in calculated.items():
                if not is_close(actual_score[field], expected):
                    score_errors.append(f"{sku}/{model}: {field}")
                    break

    audit.check(
        "Forecast evaluation is leakage-free and protocol-complete",
        not protocol_errors and len(grouped) == 560 and len(rolling) == 23520,
        "80 SKUs x 7 models x 7 origins x 6 horizons; train-only seasonal scales" if not protocol_errors else protocol_errors[0],
    )
    audit.check(
        "Forecast ME/RMSE/MAE/MASE/RMSSE/WAPE/coverage scores reconcile",
        not score_errors,
        "All 560 model-score rows independently recomputed" if not score_errors else score_errors[0],
    )
    audit.check(
        "Forecast intervals and TiRex2 quantiles are ordered and truthfully labeled",
        not interval_errors,
        "Classical 80/95% and TiRex2 native q10-q90 semantics pass" if not interval_errors else interval_errors[0],
    )

    selection_errors = []
    selected_counts: defaultdict[str, int] = defaultdict(int)
    scores_by_sku: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    for row in scores:
        scores_by_sku[row["SKU"]].append(row)
    for sku, rows in scores_by_sku.items():
        ranked = sorted(rows, key=lambda row: (float(row["RMSE"]), float(row["MAE"]), row["Model"]))
        best_rmse = float(ranked[0]["RMSE"])
        eligible = [row for row in ranked if float(row["RMSE"]) <= best_rmse * 1.02 + 1e-12]
        expected = min(eligible, key=lambda row: (MODEL_COMPLEXITY[row["Model"]], float(row["RMSE"]), float(row["MAE"]), row["Model"]))["Model"]
        selected = [row["Model"] for row in rows if bool_value(row["Selected"])]
        if selected != [expected] or any(row["Selected_Model"] != expected for row in rows):
            selection_errors.append(sku)
        selected_counts[expected] += 1
    audit.check(
        "Per-SKU model selection follows the declared 2% simplicity rule",
        not selection_errors and len(scores_by_sku) == 80,
        "Exactly one selected model per SKU" if not selection_errors else f"Mismatches: {selection_errors[:5]}",
    )

    final_grouped: defaultdict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in all_forecasts:
        final_grouped[(row["SKU"], row["Model"])].append(row)
    selected_grouped: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    for row in selected_forecasts:
        selected_grouped[row["SKU"]].append(row)
    final_ok = (
        len(final_grouped) == 560
        and all(len(rows) == 12 for rows in final_grouped.values())
        and len(selected_grouped) == 80
        and all(len(rows) == 12 for rows in selected_grouped.values())
    )
    audit.check("Every candidate and selected model has a 12-month final forecast", final_ok, f"{len(all_forecasts):,} candidate rows; {len(selected_forecasts):,} selected rows")

    audit.metrics["forecast"] = {
        "models": models,
        "rolling_prediction_rows": len(rolling),
        "score_rows": len(scores),
        "selected_models": dict(sorted(selected_counts.items())),
        "tirex2_selected_skus": selected_counts["TiRex2"],
    }


def audit_decomposition(audit: Audit) -> None:
    rows = csv_records(FORECAST_DIR / "decomposition.csv")
    features = csv_records(FORECAST_DIR / "decomposition_features.csv")
    grouped: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    identity_errors = []
    for row in rows:
        grouped[row["SKU"]].append(row)
        observed = float(row["Observed"])
        reconstructed = float(row["Trend"]) + float(row["Seasonal"]) + float(row["Remainder"])
        seasonally_adjusted = observed - float(row["Seasonal"])
        if not is_close(observed, reconstructed, abs_tol=1e-7) or not is_close(row["Seasonally_Adjusted"], seasonally_adjusted, abs_tol=1e-7):
            identity_errors.append(row["SKU"])
    feature_map = {row["SKU"]: row for row in features}
    feature_errors = []
    for sku, sku_rows in grouped.items():
        remainders = [float(row["Remainder"]) for row in sku_rows]
        detrended = [float(row["Seasonal"]) + float(row["Remainder"]) for row in sku_rows]
        deseasonalized = [float(row["Trend"]) + float(row["Remainder"]) for row in sku_rows]
        remainder_variance = statistics.pvariance(remainders)
        trend_strength = max(0.0, 1.0 - remainder_variance / statistics.pvariance(deseasonalized))
        seasonal_strength = max(0.0, 1.0 - remainder_variance / statistics.pvariance(detrended))
        feature = feature_map.get(sku)
        if (
            feature is None
            or len(sku_rows) != 36
            or integer(feature["Seasonal_Cycles"]) != 3
            or not is_close(feature["Trend_Strength"], trend_strength, abs_tol=1e-7)
            or not is_close(feature["Seasonal_Strength"], seasonal_strength, abs_tol=1e-7)
            or not is_close(feature["Remainder_SD"], statistics.stdev(remainders), abs_tol=1e-7)
        ):
            feature_errors.append(sku)
    audit.check(
        "Robust STL decomposition identities and strength features reconcile",
        not identity_errors and not feature_errors and len(rows) == 2880 and len(features) == 80,
        "Observed = trend + seasonal + remainder for 80 series" if not identity_errors and not feature_errors else f"Identity {identity_errors[:3]}; feature {feature_errors[:3]}",
    )
    audit.warn(
        "Seasonal evidence is limited",
        "Each series contains exactly three annual cycles; decomposition is diagnostic, not definitive.",
        80,
    )


def audit_manifest(audit: Audit) -> None:
    manifest = json.loads((FORECAST_DIR / "artifact_manifest.json").read_text(encoding="utf-8"))
    errors = []
    if manifest["input_sha256"] != sha256_file(WORKBOOK_PATH):
        errors.append("workbook")
    for filename, expected in manifest["artifacts"].items():
        path = FORECAST_DIR / filename
        if not path.exists() or path.stat().st_size != int(expected["bytes"]) or sha256_file(path) != expected["sha256"]:
            errors.append(filename)
    audit.check(
        "Forecast artifact manifest proves exact workbook and file provenance",
        not errors,
        f"Workbook plus {len(manifest['artifacts'])} artifacts verified" if not errors else f"Mismatch: {errors[:5]}",
    )
    audit.metrics["workbook_sha256"] = sha256_file(WORKBOOK_PATH)


def main() -> int:
    audit = Audit()
    audit_immutable_dashboards(audit)
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    required_sheets = {
        "Demand_Data", "Demand_History", "Product_Master", "DDMRP_Positions", "DDMRP_Recommendations",
        "DDMRP_Master", "Inventory_Snapshot", "Open_Supply", "Qualified_Demand", "Spend_Lines",
        "Spend_Control", "Category_Risk", "Contracts",
    }
    missing = sorted(required_sheets - set(workbook.sheetnames))
    audit.check("Workbook contract contains all required sheets", not missing, "All required sheets present" if not missing else f"Missing: {', '.join(missing)}")
    if missing:
        result = {"status": "FAIL", "checks": audit.checks, "warnings": audit.warnings, "metrics": audit.metrics}
        OUTPUT_PATH.write_text(json.dumps(result, indent=2), encoding="utf-8")
        return 1

    sheets = {name: workbook_records(workbook, name) for name in required_sheets}
    histories = audit_demand(audit, sheets)
    audit_ddmrp(audit, sheets, histories)
    audit_spend(audit, sheets)
    audit_item_segmentation(audit, sheets)
    audit_forecasts(audit, histories)
    audit_decomposition(audit)
    audit_manifest(audit)

    result = {
        "status": "FAIL" if audit.failures else "PASS_WITH_WARNINGS" if audit.warnings else "PASS",
        "generated_at_utc": datetime.now(UTC).isoformat(),
        "checks": audit.checks,
        "warnings": audit.warnings,
        "metrics": audit.metrics,
    }
    OUTPUT_PATH.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(json.dumps(result, indent=2))
    return 1 if audit.failures else 0


if __name__ == "__main__":
    sys.exit(main())
