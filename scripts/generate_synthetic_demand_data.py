"""Generate deterministic synthetic demand, DDMRP, and procurement data for Demand Genie."""

from __future__ import annotations

import calendar
import math
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_PATH = Path("data/Demand_Genie_Synthetic_Demand_History.xlsx")
SEED = 20260721
SNAPSHOT_DATE = date(2026, 7, 21)
SNAPSHOT_TIMESTAMP = "2026-07-21T06:00:00Z"
LOCATION = "PLANT-01"
MONTHS = [date(year, month, 1) for year in range(2023, 2027) for month in range(1, 13)][6:42]
SKU_MULTIPLIERS = (0.55, 0.68, 0.8, 0.95, 1.1, 1.3, 1.55, 1.85)
MAKE_GROUPS = {"GBX", "HYP", "CVL", "DRV", "CNV", "PCK"}

GROUPS = (
    {"code": "BRG", "name": "Industrial Bearings", "base": 760, "seasonality": 0.13, "trend": 0.004, "volatility": 0.06, "profile": "Seasonal", "price": 78, "lead_time": 42},
    {"code": "GBX", "name": "Gearbox Components", "base": 440, "seasonality": 0.09, "trend": 0.006, "volatility": 0.08, "profile": "Trending", "price": 215, "lead_time": 63},
    {"code": "HYP", "name": "Hydraulic Pumps", "base": 285, "seasonality": 0.16, "trend": 0.003, "volatility": 0.12, "profile": "Volatile", "price": 540, "lead_time": 70},
    {"code": "SEA", "name": "Sealing Systems", "base": 1220, "seasonality": 0.07, "trend": 0.002, "volatility": 0.05, "profile": "Stable", "price": 29, "lead_time": 28},
    {"code": "CVL", "name": "Control Valves", "base": 330, "seasonality": 0.11, "trend": -0.002, "volatility": 0.1, "profile": "Volatile", "price": 310, "lead_time": 56},
    {"code": "DRV", "name": "Electric Drives", "base": 245, "seasonality": 0.14, "trend": 0.007, "volatility": 0.09, "profile": "Trending", "price": 680, "lead_time": 77},
    {"code": "SNS", "name": "Sensor Modules", "base": 510, "seasonality": 0.05, "trend": 0.01, "volatility": 0.13, "profile": "Growth", "price": 168, "lead_time": 49},
    {"code": "CNV", "name": "Conveying Components", "base": 620, "seasonality": 0.17, "trend": 0.001, "volatility": 0.08, "profile": "Seasonal", "price": 115, "lead_time": 45},
    {"code": "FST", "name": "Fastening Kits", "base": 1880, "seasonality": 0.06, "trend": 0.002, "volatility": 0.07, "profile": "Stable", "price": 16, "lead_time": 21},
    {"code": "PCK", "name": "Packaging Equipment", "base": 165, "seasonality": 0.2, "trend": 0.005, "volatility": 0.15, "profile": "Intermittent", "price": 950, "lead_time": 84},
)

DESCRIPTORS = ("Core", "Compact", "Standard", "Heavy Duty", "Precision", "High Capacity", "Service Kit", "Premium")
PLANNERS = ("Alex Planner", "Sam Buyer", "Jordan Scheduler", "Taylor Planner")
BUYERS = ("Sam Buyer", "Morgan Buyer", "Casey Sourcing", "Riley Category")

IMPACT_WEIGHTS = (0.30, 0.30, 0.20, 0.10, 0.10)
RISK_WEIGHTS = (0.20, 0.15, 0.15, 0.15, 0.10, 0.10, 0.10, 0.05)
CATEGORY_PROFILES = {
    "BRG": {"family": "Mechanical Components", "impact": (75, 72, 64, 50, 45), "risk": (32, 45, 44, 40, 35, 28, 42, 36), "supplier_count": 4, "weights": (0.53, 0.25, 0.14, 0.08)},
    "GBX": {"family": "Mechanical Components", "impact": (85, 92, 86, 70, 75), "risk": (72, 78, 76, 80, 52, 50, 68, 42), "supplier_count": 3, "weights": (0.57, 0.29, 0.14)},
    "HYP": {"family": "Fluid Power & Control", "impact": (88, 95, 90, 84, 72), "risk": (82, 86, 80, 74, 64, 58, 72, 55), "supplier_count": 2, "weights": (0.68, 0.32)},
    "SEA": {"family": "Mechanical Components", "impact": (42, 38, 30, 35, 24), "risk": (24, 20, 28, 22, 18, 22, 26, 30), "supplier_count": 6, "weights": (0.25, 0.22, 0.18, 0.15, 0.12, 0.08)},
    "CVL": {"family": "Fluid Power & Control", "impact": (76, 84, 78, 72, 66), "risk": (61, 68, 64, 58, 44, 38, 58, 42), "supplier_count": 4, "weights": (0.48, 0.28, 0.16, 0.08)},
    "DRV": {"family": "Electrical & Automation", "impact": (94, 96, 92, 78, 88), "risk": (78, 82, 88, 80, 58, 42, 64, 48), "supplier_count": 3, "weights": (0.61, 0.27, 0.12)},
    "SNS": {"family": "Electrical & Automation", "impact": (44, 48, 40, 62, 72), "risk": (88, 90, 76, 84, 72, 54, 70, 52), "supplier_count": 2, "weights": (0.86, 0.14)},
    "CNV": {"family": "Material Handling", "impact": (64, 70, 55, 45, 52), "risk": (38, 42, 44, 36, 32, 30, 38, 35), "supplier_count": 5, "weights": (0.34, 0.25, 0.19, 0.13, 0.09)},
    "FST": {"family": "Mechanical Components", "impact": (31, 28, 24, 20, 18), "risk": (18, 16, 22, 20, 14, 16, 20, 25), "supplier_count": 7, "weights": (0.22, 0.19, 0.17, 0.14, 0.12, 0.09, 0.07)},
    "PCK": {"family": "Material Handling", "impact": (46, 48, 40, 36, 44), "risk": (72, 78, 86, 70, 64, 48, 62, 44), "supplier_count": 2, "weights": (0.84, 0.16)},
}

SUPPLIER_ROOTS = (
    "Asteron", "Norden", "Helix", "Vantage", "Alpine", "Kinetic", "Meridian", "Solstice", "Vertex", "Cobalt",
    "Rheinwerk", "Orion", "DeltaForge", "Lumina", "Axis", "Hartmann", "Novatek", "Keystone", "Boreal", "Crestline",
    "Vectora", "Summit", "Precisiona", "Coreline", "Dynatek", "Bridgeway", "Fortis", "Arcwell", "Evergreen", "Pinnacle",
    "Westmark", "Northstar", "Axiom", "Sterling", "Redwood", "BluePeak", "Ironclad", "Clearpath", "Frontier", "Unity",
)
SUPPLIER_SUFFIXES = ("Industries", "Components", "Systems", "Technologies", "Manufacturing")
SUPPLIER_COUNTRIES = (
    ("Germany", "Europe", "EUR", "DAP"),
    ("Czech Republic", "Europe", "EUR", "FCA"),
    ("United States", "Americas", "USD", "DAP"),
    ("United Kingdom", "Europe", "GBP", "FCA"),
    ("Switzerland", "Europe", "CHF", "DAP"),
    ("China", "Asia Pacific", "CNY", "FOB"),
    ("Poland", "Europe", "EUR", "FCA"),
    ("Mexico", "Americas", "USD", "FCA"),
)
FX_TO_EUR = {"EUR": 1.0, "USD": 0.92, "GBP": 1.17, "CHF": 1.04, "CNY": 0.128}

HEADER_FILL = PatternFill("solid", fgColor="1F5E69")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCEBED")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(bold=True, color="17262B")
THIN_LINE = Side(style="thin", color="CBD6D7")


def calculate_demand(group: dict[str, float | int | str], sku_index: int, month_index: int, rng: random.Random) -> int:
    """Create a positive monthly demand value with useful forecasting characteristics."""
    multiplier = SKU_MULTIPLIERS[sku_index]
    phase = (sku_index * 0.61) + (len(str(group["code"])) * 0.2)
    seasonality = 1 + float(group["seasonality"]) * math.sin((month_index / 12 * 2 * math.pi) + phase)
    trend = 1 + float(group["trend"]) * month_index
    noise = 1 + rng.gauss(0, float(group["volatility"]))
    shock = 1.0

    if (sku_index + month_index * 3) % 29 == 0:
        shock *= 1.16
    if (sku_index * 5 + month_index) % 37 == 0:
        shock *= 0.82
    if group["profile"] == "Intermittent" and (month_index + sku_index) % 9 == 0:
        shock *= 0.32
    if group["profile"] == "Growth" and month_index >= 24:
        shock *= 1.08

    demand = float(group["base"]) * multiplier * seasonality * trend * noise * shock
    return max(1, int(round(demand)))


def make_sku(group: dict[str, float | int | str], sku_index: int) -> dict[str, str | float | int]:
    sku = f"{group['code']}-{sku_index + 1:02d}"
    group_code = str(group["code"])
    supply_type = "Make" if group_code in MAKE_GROUPS else "Buy"
    category = CATEGORY_PROFILES[group_code]
    impact_score, _ = category_scores(group_code)
    criticality = "Critical" if impact_score >= 75 and sku_index >= 4 else "High" if impact_score >= 60 or sku_index >= 6 else "Medium" if impact_score >= 35 else "Low"
    lifecycle = "New Introduction" if group_code == "SNS" and sku_index >= 6 else "Phase-out" if group_code == "CVL" and sku_index <= 1 else "Active"
    unit_price = round(float(group["price"]) * (0.78 + sku_index * 0.07), 2)
    unit_cost = round(unit_price * (0.54 + (sum(ord(char) for char in group_code) % 11) / 100), 2)
    source_count = min(int(category["supplier_count"]), 1 + (sku_index % 4))
    return {
        "product_group_code": group_code,
        "product_group": str(group["name"]),
        "sku": sku,
        "description": f"{group['name']} {DESCRIPTORS[sku_index]}",
        "profile": str(group["profile"]),
        "unit_price": unit_price,
        "unit_cost": unit_cost,
        "lead_time": int(group["lead_time"]) + (sku_index % 4) * 3,
        "abc_class": "A" if sku_index >= 5 else "B" if sku_index >= 2 else "C",
        "xyz_class": "X" if group["profile"] in {"Stable", "Growth"} else "Y" if group["profile"] in {"Seasonal", "Trending"} else "Z",
        "supply_type": supply_type,
        "criticality": criticality,
        "criticality_score": {"Critical": 95, "High": 75, "Medium": 50, "Low": 25}[criticality],
        "lifecycle": lifecycle,
        "source_count": source_count,
        "single_source": "Yes" if source_count == 1 else "No",
        "substitute_count": 0 if source_count == 1 else 1 if source_count == 2 else 2,
        "revenue_at_risk": round(float(group["base"]) * SKU_MULTIPLIERS[sku_index] * 12 * unit_price * (0.18 + sku_index * 0.025), 2),
        "spend_category_code": group_code,
        "sku_index": sku_index,
    }


def weighted_average(values: tuple[int, ...], weights: tuple[float, ...]) -> float:
    return round(sum(value * weight for value, weight in zip(values, weights)), 1)


def category_scores(group_code: str) -> tuple[float, float]:
    profile = CATEGORY_PROFILES[group_code]
    return weighted_average(profile["impact"], IMPACT_WEIGHTS), weighted_average(profile["risk"], RISK_WEIGHTS)


def kraljic_quadrant(impact_score: float, risk_score: float) -> str:
    if impact_score >= 50 and risk_score >= 50:
        return "Strategic"
    if impact_score >= 50:
        return "Leverage"
    if risk_score >= 50:
        return "Bottleneck"
    return "Routine"


def round_up(value: float, multiple: int) -> int:
    return int(math.ceil(value / multiple) * multiple) if value > 0 else 0


def days_in_month(month: date) -> int:
    return calendar.monthrange(month.year, month.month)[1]


def dlt_category(dlt_days: int) -> str:
    if dlt_days >= 60:
        return "Long"
    if dlt_days >= 35:
        return "Medium"
    return "Short"


def lead_time_factor(dlt_days: int) -> float:
    return 0.3 if dlt_days >= 60 else 0.5 if dlt_days >= 35 else 0.75


def variability_factor(xyz_class: str) -> float:
    return {"X": 0.25, "Y": 0.5, "Z": 0.75}[xyz_class]


def nfp_status(net_flow: int, top_red: float, top_yellow: float, top_green: float) -> str:
    if net_flow < 0:
        return "Black"
    if net_flow <= top_red:
        return "Red"
    if net_flow <= top_yellow:
        return "Yellow"
    if net_flow <= top_green:
        return "Green"
    return "Blue"


def planning_action(status: str) -> str:
    if status == "Black":
        return "Critical: release supply recommendation and investigate immediate flow risk"
    if status == "Red":
        return "Urgent: release supply recommendation and review execution risk"
    if status == "Yellow":
        return "Replenish: release supply recommendation"
    if status == "Blue":
        return "No routine replenishment: review excess before changing open supply"
    return "No routine replenishment"


def style_sheet(sheet, widths: dict[str, float], freeze: str = "A2") -> None:
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="left")
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def add_table(sheet, name: str) -> None:
    table = Table(displayName=name, ref=f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    sheet.add_table(table)


def format_columns(sheet, *, dates: tuple[str, ...] = (), integers: tuple[str, ...] = (), decimals: tuple[str, ...] = ()) -> None:
    header_columns = {cell.value: cell.column for cell in sheet[1]}
    for header in dates:
        if header in header_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, header_columns[header]).number_format = "yyyy-mm-dd"
    for header in integers:
        if header in header_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, header_columns[header]).number_format = "#,##0"
    for header in decimals:
        if header in header_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, header_columns[header]).number_format = "#,##0.00"


def format_currency_columns(sheet, headers: tuple[str, ...]) -> None:
    header_columns = {cell.value: cell.column for cell in sheet[1]}
    for header in headers:
        if header in header_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, header_columns[header]).number_format = '#,##0.00 "EUR"'


def build_ddmrp_rows(
    all_skus: list[dict[str, str | float | int]], histories: dict[str, list[int]], rng: random.Random
) -> dict[str, list[list[object]]]:
    """Build item-location DDMRP master, operational source, and recommendation records."""
    history_days = sum(days_in_month(month) for month in MONTHS[-12:])
    output = {
        "decoupling": [],
        "master": [],
        "inventory": [],
        "supply": [],
        "qualified_demand": [],
        "positions": [],
        "recommendations": [],
        "adjustments": [],
    }

    for position_index, sku_data in enumerate(all_skus):
        sku = str(sku_data["sku"])
        history = histories[sku]
        supply_type = str(sku_data["supply_type"])
        nominal_lead_time = int(sku_data["lead_time"])
        dlt_days = nominal_lead_time if supply_type == "Buy" else max(7, int(round(nominal_lead_time * 0.65)))
        baseline_adu = sum(history[-12:]) / history_days
        ltf = lead_time_factor(dlt_days)
        vf = variability_factor(str(sku_data["xyz_class"]))
        order_cycle = 14 if supply_type == "Buy" else 7
        order_multiple = 50 if baseline_adu >= 50 else 25 if baseline_adu >= 20 else 10 if baseline_adu >= 5 else 1
        moq = round_up(baseline_adu * order_cycle * 0.8, order_multiple)
        spike_horizon = 21 if dlt_days >= 60 else 14
        spike_threshold = round_up(max(baseline_adu * 1.8, order_multiple), order_multiple)
        adjustment_factor = 1.0
        adjustment_reason = "No planned adjustment"
        adjustment_status = "None"
        adjustment_from = ""
        adjustment_through = ""

        if str(sku_data["product_group_code"]) in {"CNV", "PCK"} and position_index % 2 == 0:
            adjustment_factor = 1.15
            adjustment_reason = "Synthetic approved seasonal event for dashboard testing"
            adjustment_status = "Approved synthetic demo"
            adjustment_from = SNAPSHOT_DATE
            adjustment_through = date(2026, 9, 30)
        elif str(sku_data["product_group_code"]) == "BRG" and position_index % 7 == 0:
            adjustment_factor = 0.92
            adjustment_reason = "Synthetic approved maintenance-period adjustment"
            adjustment_status = "Approved synthetic demo"
            adjustment_from = SNAPSHOT_DATE
            adjustment_through = date(2026, 8, 31)

        adjusted_adu = baseline_adu * adjustment_factor
        yellow = adjusted_adu * dlt_days
        red_base = yellow * ltf
        red_safety = red_base * vf
        red = red_base + red_safety
        green = max(moq, adjusted_adu * order_cycle, adjusted_adu * dlt_days * ltf)
        top_red = red
        top_yellow = red + yellow
        top_green = top_yellow + green
        planner = PLANNERS[position_index % len(PLANNERS)]
        decoupling_reason = (
            "Supplier variability and customer lead-time protection"
            if supply_type == "Buy"
            else "Critical operation protection with upstream decoupling"
        )
        dlt_basis = "Supplier lead time at external source" if supply_type == "Buy" else "Synthetic BOM/routing path to upstream buffer"
        upstream_source = "Approved supplier schedule" if supply_type == "Buy" else "Upstream component buffer and routing"
        buffer_profile = f"{supply_type[0]}-{dlt_category(dlt_days)[0]}-{sku_data['xyz_class']}"

        today_demand = max(1, int(round(baseline_adu * rng.uniform(0.75, 1.3))))
        past_due_demand = int(round(baseline_adu * rng.uniform(0.2, 0.8))) if position_index % 4 == 0 else 0
        qualified_spike_demand = max(spike_threshold, int(round(baseline_adu * 2.4))) if position_index % 6 == 0 else 0
        qualified_demand = past_due_demand + today_demand + qualified_spike_demand

        status_pattern = position_index % 5
        if status_pattern == 0:
            target_nfp = -min(max(1, qualified_demand * 0.45), top_green * 0.08)
        elif status_pattern == 1:
            target_nfp = top_red * 0.6
        elif status_pattern == 2:
            target_nfp = top_red + yellow * 0.55
        elif status_pattern == 3:
            target_nfp = top_yellow + green * 0.5
        else:
            target_nfp = top_green * 1.15

        total_available = max(0, int(round(target_nfp + qualified_demand)))
        if position_index % 9 == 0:
            on_hand = min(total_available, max(0, int(round(top_red * 0.4))))
        else:
            on_hand_ratio = (0.4, 0.5, 0.6, 0.75, 0.85)[status_pattern]
            on_hand = int(round(total_available * on_hand_ratio))
        open_supply = max(0, total_available - on_hand)
        net_flow = on_hand + open_supply - qualified_demand
        status = nfp_status(net_flow, top_red, top_yellow, top_green)
        triggered = net_flow <= top_yellow
        base_recommendation = max(top_green - net_flow, 0) if triggered else 0
        recommended_order = round_up(base_recommendation, order_multiple)
        execution_review = "Review open supply timing and synchronization: on hand is at or below TOR" if on_hand <= top_red else "No automatic on-hand red-zone review"

        output["decoupling"].append([
            SNAPSHOT_DATE, LOCATION, sku, sku_data["description"], supply_type, "Approved", dlt_days, dlt_basis,
            upstream_source, decoupling_reason, planner, date(2026, 6, 30), date(2026, 10, 1), "Synthetic approved demo",
        ])
        output["master"].append([
            SNAPSHOT_DATE, LOCATION, sku, sku_data["product_group"], sku_data["description"], buffer_profile, supply_type,
            baseline_adu, "Trailing 12 complete months / calendar days", dlt_days, ltf, vf, moq, order_cycle,
            order_multiple, spike_horizon, spike_threshold, adjustment_factor, adjustment_status, planner,
            "Approved synthetic demo", date(2026, 10, 1),
        ])
        output["inventory"].append([
            SNAPSHOT_DATE, SNAPSHOT_TIMESTAMP, LOCATION, sku, sku_data["description"], on_hand, 0, on_hand,
            "Available", "Complete synthetic inventory snapshot",
        ])
        output["adjustments"].append([
            LOCATION, sku, SNAPSHOT_DATE, adjustment_from, adjustment_through, adjustment_factor, adjustment_status,
            adjustment_reason, "Synthetic event data; replace with FPP3 forecast proposal before operational use", planner,
        ])

        if past_due_demand > 0:
            output["qualified_demand"].append([
                f"SO-PD-{sku}", LOCATION, sku, SNAPSHOT_DATE - timedelta(days=1), "Past_Due", past_due_demand,
                "Yes", "Past-due sales order demand", spike_threshold,
            ])
        output["qualified_demand"].append([
            f"SO-TD-{sku}", LOCATION, sku, SNAPSHOT_DATE, "Today", today_demand, "Yes", "Sales order demand due today", spike_threshold,
        ])
        if qualified_spike_demand > 0:
            output["qualified_demand"].append([
                f"SO-SP-{sku}", LOCATION, sku, SNAPSHOT_DATE + timedelta(days=3 + (position_index % 8)), "Qualified_Spike",
                qualified_spike_demand, "Yes", "Future order exceeds approved spike threshold inside spike horizon", spike_threshold,
            ])

        if open_supply > 0:
            split_supply = position_index % 3 == 0 and open_supply > order_multiple * 2
            first_supply = int(round(open_supply * 0.6)) if split_supply else open_supply
            second_supply = open_supply - first_supply
            first_status = "Late" if position_index % 9 == 0 else "Firmed"
            first_receipt = SNAPSHOT_DATE - timedelta(days=2) if first_status == "Late" else SNAPSHOT_DATE + timedelta(days=max(1, int(round(dlt_days * 0.6))))
            output["supply"].append([
                f"{'PO' if supply_type == 'Buy' else 'MO'}-{sku}-01", LOCATION, sku, supply_type, first_supply,
                first_receipt, first_status, "Yes", "Supplier A" if supply_type == "Buy" else "Assembly Cell 1",
                "Late receipt requires execution review" if first_status == "Late" else "No current synchronization exception",
            ])
            if second_supply > 0:
                output["supply"].append([
                    f"{'PO' if supply_type == 'Buy' else 'MO'}-{sku}-02", LOCATION, sku, supply_type, second_supply,
                    SNAPSHOT_DATE + timedelta(days=max(1, int(round(dlt_days * 0.9)))), "Released", "Yes",
                    "Supplier B" if supply_type == "Buy" else "Assembly Cell 2", "No current synchronization exception",
                ])

        output["positions"].append([
            sku, LOCATION, sku_data["description"], sku_data["product_group"], baseline_adu, dlt_days, ltf, vf, moq,
            order_cycle, on_hand, open_supply, past_due_demand, today_demand, qualified_spike_demand, order_multiple,
            adjustment_factor, SNAPSHOT_TIMESTAMP,
        ])
        output["recommendations"].append([
            SNAPSHOT_DATE, LOCATION, sku, sku_data["description"], sku_data["product_group"], buffer_profile, baseline_adu,
            adjustment_factor, adjusted_adu, dlt_days, ltf, vf, moq, order_cycle, order_multiple, red_base, red_safety,
            red, yellow, green, top_red, top_yellow, top_green, on_hand, open_supply, past_due_demand, today_demand,
            qualified_spike_demand, qualified_demand, net_flow, net_flow / top_green * 100 if top_green else 0, status,
            "Yes" if triggered else "No", base_recommendation, recommended_order,
            SNAPSHOT_DATE + timedelta(days=dlt_days), planning_action(status), on_hand - top_red, execution_review,
            "Proposed - planner approval required", "Complete synthetic DDMRP input set",
        ])

    return output


def build_procurement_rows(
    all_skus: list[dict[str, str | float | int]], histories: dict[str, list[int]]
) -> dict[str, list[dict[str, object]]]:
    """Build a separate, signed procurement ledger and its commercial master data."""
    rng = random.Random(SEED + 9107)
    group_lookup = {str(group["code"]): group for group in GROUPS}
    supplier_map: dict[str, list[dict[str, object]]] = {}
    suppliers: list[dict[str, object]] = []
    contracts: list[dict[str, object]] = []
    supplier_number = 0

    for group_index, group in enumerate(GROUPS):
        group_code = str(group["code"])
        category = CATEGORY_PROFILES[group_code]
        group_suppliers = []
        for supplier_index in range(int(category["supplier_count"])):
            root = SUPPLIER_ROOTS[supplier_number]
            suffix = SUPPLIER_SUFFIXES[(group_index + supplier_index) % len(SUPPLIER_SUFFIXES)]
            supplier_id = f"SUP-{group_code}-{supplier_index + 1:02d}"
            parent_id = f"PAR-{group_code}-{supplier_index + 1:02d}"
            country, region, currency, incoterm = SUPPLIER_COUNTRIES[(supplier_number + group_index) % len(SUPPLIER_COUNTRIES)]
            preferred = supplier_index < 2
            contract_id = f"CTR-{group_code}-{supplier_index + 1:02d}" if preferred else ""
            normalized_name = f"{root} {suffix}"
            risk_inputs = category["risk"]
            supplier = {
                "Supplier_ID": supplier_id,
                "Supplier_Normalized_ID": supplier_id,
                "Supplier_Normalized_Name": normalized_name,
                "Supplier_Parent_ID": parent_id,
                "Supplier_Parent_Name": normalized_name,
                "Category_Code": group_code,
                "Category_L1": str(group["name"]),
                "Country": country,
                "Region": region,
                "Original_Currency": currency,
                "Preferred_Supplier": "Yes" if preferred else "No",
                "Approved_Status": "Approved",
                "Supplier_Tier": "Strategic" if preferred and weighted_average(category["impact"], IMPACT_WEIGHTS) >= 70 else "Approved",
                "Financial_Risk_Score": max(5, min(95, int(risk_inputs[5]) + supplier_index * 3 - 2)),
                "Quality_Risk_Score": max(5, min(95, int(risk_inputs[6]) + supplier_index * 2 - 1)),
                "Delivery_Risk_Score": max(5, min(95, int(risk_inputs[3]) + supplier_index * 4)),
                "ESG_Risk_Score": max(5, min(95, int(risk_inputs[7]) + supplier_index * 2)),
                "Payment_Terms_Days": 45 if preferred else 30,
                "Default_Incoterm": incoterm,
                "Default_Contract_ID": contract_id,
                "Risk_As_Of": SNAPSHOT_DATE,
                "Risk_Owner": BUYERS[group_index % len(BUYERS)],
                "Data_Status": "Complete synthetic supplier record",
                "supplier_index": supplier_index,
            }
            suppliers.append(supplier)
            group_suppliers.append(supplier)
            if preferred:
                contracts.append({
                    "Contract_ID": contract_id,
                    "Supplier_Normalized_ID": supplier_id,
                    "Supplier_Normalized_Name": normalized_name,
                    "Category_Code": group_code,
                    "Category_L1": str(group["name"]),
                    "Contract_Status": "Active",
                    "Start_Date": date(2024, 7, 1),
                    "End_Date": date(2027, 6, 30),
                    "Contract_Owner": BUYERS[group_index % len(BUYERS)],
                    "Payment_Terms_Days": supplier["Payment_Terms_Days"],
                    "Rebate_Percent": 2.5 if supplier_index == 0 else 1.5,
                    "Price_Adjustment_Clause": "Annual index review with documented approval",
                    "Committed_Annual_Value_EUR": round(float(group["base"]) * 12 * float(group["price"]) * (0.42 if supplier_index == 0 else 0.2), 2),
                    "Renewal_Action": "Review by 2027-03-31",
                    "Data_Status": "Complete synthetic contract record",
                })
            supplier_number += 1
        supplier_map[group_code] = group_suppliers

    spend_lines: list[dict[str, object]] = []
    procurement_months = MONTHS[-24:]
    leakage_rates = {"CVL": 0.28, "DRV": 0.24, "PCK": 0.31, "SNS": 0.18}
    line_counter = 0

    for sku_position, sku_data in enumerate(all_skus):
        sku = str(sku_data["sku"])
        group_code = str(sku_data["product_group_code"])
        group = group_lookup[group_code]
        category = CATEGORY_PROFILES[group_code]
        impact_score, risk_score = category_scores(group_code)
        category_suppliers = supplier_map[group_code]
        for month_index, month in enumerate(procurement_months):
            history_index = len(MONTHS) - len(procurement_months) + month_index
            demand_units = histories[sku][history_index]
            order_multiple = 5 if demand_units < 100 else 10 if demand_units < 500 else 25
            purchase_factor = 0.94 + ((sku_position + month_index) % 9) * 0.015
            quantity = max(order_multiple, round_up(demand_units * purchase_factor, order_multiple))
            supplier = rng.choices(category_suppliers, weights=category["weights"], k=1)[0]
            supplier_index = int(supplier["supplier_index"])
            preferred = supplier["Preferred_Supplier"] == "Yes"
            leakage = leakage_rates.get(group_code, 0.08)
            on_contract = preferred and rng.random() >= leakage
            contract_id = str(supplier["Default_Contract_ID"]) if on_contract else ""
            transaction_date = month.replace(day=5 + ((month_index * 3 + sku_position) % 20))
            po_date = transaction_date - timedelta(days=12 + (sku_position % 9))
            currency = str(supplier["Original_Currency"])
            fx_rate = 1.0 if currency == "EUR" else round(FX_TO_EUR[currency] * (1 + 0.018 * math.sin(month_index / 3 + supplier_index)), 6)
            escalation = 1 + month_index * 0.0017
            unit_cost = round(float(sku_data["unit_cost"]) * escalation * (0.975 + supplier_index * 0.018 + rng.uniform(-0.012, 0.012)), 4)
            extended = quantity * unit_cost
            freight = round(extended * (0.012 if str(supplier["Region"]) != "Europe" else 0.0035), 2)
            discount = round(extended * (0.025 if on_contract else 0), 2)
            spend_base = round(extended + freight - discount, 2)
            invoice_id = f"INV-{transaction_date:%Y%m}-{sku.replace('-', '')}"
            po_id = f"PO-{po_date:%Y%m}-{sku.replace('-', '')}"
            raw_name_variants = (
                str(supplier["Supplier_Normalized_Name"]),
                str(supplier["Supplier_Normalized_Name"]).upper(),
                f"{supplier['Supplier_Normalized_Name']} GmbH",
            )
            raw_name = raw_name_variants[(sku_position + month_index) % len(raw_name_variants)]
            line_counter += 1
            common = {
                "Source_System": "SYNTH_ERP_AP",
                "Transaction_ID": invoice_id,
                "Line_ID": "1",
                "Transaction_Date": transaction_date,
                "Supplier_ID": str(supplier["Supplier_ID"]),
                "Supplier_Name": raw_name,
                "Supplier_Normalized_ID": str(supplier["Supplier_Normalized_ID"]),
                "Supplier_Normalized_Name": str(supplier["Supplier_Normalized_Name"]),
                "Supplier_Parent_ID": str(supplier["Supplier_Parent_ID"]),
                "Supplier_Parent_Name": str(supplier["Supplier_Parent_Name"]),
                "Category_L1": str(group["name"]),
                "Category_L2": str(category["family"]),
                "Category_L3": str(group["name"]),
                "Category_Code": group_code,
                "Description_Raw": str(sku_data["description"]),
                "Description_Normalized": str(sku_data["description"]),
                "Part_ID": sku,
                "Organization": "Demand Genie Manufacturing GmbH",
                "Plant": LOCATION,
                "Cost_Center": f"CC-{group_code}-100",
                "Buyer": BUYERS[list(group_lookup).index(group_code) % len(BUYERS)],
                "PO_ID": po_id,
                "PO_Date": po_date,
                "Invoice_ID": invoice_id,
                "Contract_ID": contract_id,
                "On_Contract": "Yes" if on_contract else "No",
                "Preferred_Supplier": "Yes" if preferred else "No",
                "Quantity": quantity,
                "UOM": "EA",
                "Quantity_Base_UOM": quantity,
                "Unit_Price_Base": unit_cost,
                "Spend_Base": spend_base,
                "Base_Currency": "EUR",
                "Spend_Original": round(spend_base / fx_rate, 2),
                "Original_Currency": currency,
                "FX_Rate_to_EUR": fx_rate,
                "FX_Date": transaction_date,
                "Tax_Base": 0,
                "Freight_Base": freight,
                "Discount_Base": discount,
                "Payment_Terms_Days": int(supplier["Payment_Terms_Days"]),
                "Incoterm": str(supplier["Default_Incoterm"]),
                "Direct_Indirect": "Direct",
                "CapEx_OpEx": "OpEx",
                "Classification_Method": "Rule: product group to approved category taxonomy",
                "Classification_Confidence": 1.0,
                "Classification_Override": "No",
                "Business_Impact_Score": impact_score,
                "Supply_Risk_Score": risk_score,
                "Data_Status": "Complete synthetic procurement line",
            }
            spend_lines.append(common)

            if (sku_position * 7 + month_index * 5) % 43 == 0:
                credit = dict(common)
                credit_amount = -round(spend_base * (0.018 + (sku_position % 4) * 0.006), 2)
                credit_date = transaction_date + timedelta(days=2)
                credit_id = f"CRN-{transaction_date:%Y%m}-{sku.replace('-', '')}"
                credit.update({
                    "Transaction_ID": credit_id,
                    "Transaction_Date": credit_date,
                    "Quantity": 0,
                    "Quantity_Base_UOM": 0,
                    "Unit_Price_Base": 0,
                    "Spend_Base": credit_amount,
                    "Spend_Original": round(credit_amount / fx_rate, 2),
                    "FX_Date": credit_date,
                    "Tax_Base": 0,
                    "Freight_Base": 0,
                    "Discount_Base": 0,
                    "Invoice_ID": credit_id,
                    "Description_Raw": f"Credit correction: {sku_data['description']}",
                    "Data_Status": "Complete synthetic credit line",
                })
                spend_lines.append(credit)

    demand_history: list[dict[str, object]] = []
    for sku_data in all_skus:
        sku = str(sku_data["sku"])
        for month, demand_units in zip(MONTHS, histories[sku]):
            demand_history.append({
                "Period": month,
                "Part_ID": sku,
                "Part_Description": str(sku_data["description"]),
                "Product_Group": str(sku_data["product_group"]),
                "Location": LOCATION,
                "Demand_Units": demand_units,
                "Unit_Cost_Base": float(sku_data["unit_cost"]),
                "Base_Currency": "EUR",
                "Criticality": str(sku_data["criticality"]),
                "Criticality_Score": int(sku_data["criticality_score"]),
                "Lifecycle_Status": str(sku_data["lifecycle"]),
                "Lead_Time_Days": int(sku_data["lead_time"]),
                "Source_Count": int(sku_data["source_count"]),
                "Data_Quality_Status": "Complete synthetic monthly history",
            })

    taxonomy: list[dict[str, object]] = []
    category_risk: list[dict[str, object]] = []
    for group in GROUPS:
        group_code = str(group["code"])
        profile = CATEGORY_PROFILES[group_code]
        impact_score, risk_score = category_scores(group_code)
        impact = profile["impact"]
        risk = profile["risk"]
        taxonomy.append({
            "Category_Code": group_code,
            "Category_L1": str(group["name"]),
            "Category_L2": str(profile["family"]),
            "Category_L3": str(group["name"]),
            "Direct_Indirect": "Direct",
            "Category_Owner": BUYERS[list(group_lookup).index(group_code) % len(BUYERS)],
            "Taxonomy_Status": "Approved synthetic demo",
            "Effective_From": date(2024, 7, 1),
            "Next_Review_Date": date(2026, 10, 1),
        })
        category_risk.append({
            "Review_Date": SNAPSHOT_DATE,
            "Category_Code": group_code,
            "Category_L1": str(group["name"]),
            "Spend_Impact_Score": impact[0],
            "Operational_Criticality_Score": impact[1],
            "Revenue_at_Risk_Score": impact[2],
            "Quality_Regulatory_Score": impact[3],
            "Innovation_Impact_Score": impact[4],
            "Source_Scarcity_Score": risk[0],
            "Switching_Time_Score": risk[1],
            "Lead_Time_Risk_Score": risk[2],
            "Capacity_Risk_Score": risk[3],
            "Country_Logistics_Risk_Score": risk[4],
            "Financial_Risk_Score": risk[5],
            "Quality_Risk_Score": risk[6],
            "ESG_Risk_Score": risk[7],
            "Business_Impact_Score": impact_score,
            "Supply_Risk_Score": risk_score,
            "Kraljic_Quadrant": kraljic_quadrant(impact_score, risk_score),
            "Qualified_Supplier_Count": int(profile["supplier_count"]),
            "Switching_Time_Days": int(group["lead_time"]) * 2,
            "Category_Risk_Owner": BUYERS[list(group_lookup).index(group_code) % len(BUYERS)],
            "Evidence_Status": "Complete synthetic factor evidence",
            "Next_Review_Date": date(2026, 10, 1),
        })

    net_spend = round(sum(float(row["Spend_Base"]) for row in spend_lines), 2)
    gross_spend = round(sum(max(0, float(row["Spend_Base"])) for row in spend_lines), 2)
    credits = round(sum(min(0, float(row["Spend_Base"])) for row in spend_lines), 2)
    controls = [{
        "Control_ID": "AP-LEDGER-2024-07_TO_2026-06",
        "Scope_Start": min(row["Transaction_Date"] for row in spend_lines),
        "Scope_End": max(row["Transaction_Date"] for row in spend_lines),
        "Organization": "Demand Genie Manufacturing GmbH",
        "Base_Currency": "EUR",
        "Control_Net_Spend": net_spend,
        "Analyzed_Net_Spend": net_spend,
        "Difference": 0,
        "Gross_Positive_Spend": gross_spend,
        "Credits": credits,
        "Line_Count": len(spend_lines),
        "Control_Status": "Reconciled",
        "Control_Note": "Synthetic AP ledger control total; signed credits retained",
    }]
    return {
        "suppliers": suppliers,
        "contracts": contracts,
        "spend_lines": spend_lines,
        "demand_history": demand_history,
        "taxonomy": taxonomy,
        "category_risk": category_risk,
        "controls": controls,
    }


def append_dict_rows(sheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def build_workbook() -> None:
    rng = random.Random(SEED)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    read_me = workbook.active
    read_me.title = "Read_Me"
    read_me.append(["Demand Genie Synthetic Planning and Procurement Data"])
    read_me["A1"].fill = HEADER_FILL
    read_me["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    read_me.merge_cells("A1:D1")
    read_me.append([])
    read_me.append(["Purpose", "Synthetic monthly demand, complete DDMRP planning data, and an auditable procurement spend ledger for dashboard development. All values are synthetic demo data, not live ERP or finance data."])
    read_me.append(["Coverage", "36 complete months of demand: July 2023 through June 2026; 24 months of procurement: July 2024 through June 2026; DDMRP planning snapshot: 2026-07-21 at PLANT-01."])
    read_me.append(["Scope", "10 product groups, 8 SKUs per group, 80 total SKUs, 2,880 monthly demand records, 80 calculator-ready DDMRP positions, and a signed line-level procurement ledger."])
    read_me.append(["Forecasting", "Use Demand_Data for FPP3 forecast training. Forecast output should later create approval-required factors in DDMRP_Adjustments; it must not be inserted directly into qualified demand."])
    read_me.append(["DDMRP planning", "DDMRP_Positions contains the complete required input grain for the DDMRP calculator. DDMRP_Recommendations is the transparent synthetic output for the snapshot."])
    read_me.append(["Data separation", "Inventory_Snapshot, Open_Supply, and Qualified_Demand are source-level operational data. DDMRP_Master and DDMRP_Decoupling contain approved synthetic configuration and DLT design."])
    read_me.append(["Data note", "Demand includes trend, seasonality, volatility, shocks, and intermittent behavior. DDMRP positions intentionally include black, red, yellow, green, and blue net-flow states for dashboard testing."])
    read_me.append(["Spend analysis", "Use Spend_Lines for procurement analysis. Gross positive spend drives Pareto ranking; signed net spend reconciles to Spend_Control; credits remain visible and are never discarded."])
    read_me.append(["Portfolio method", "Category_Risk contains explicit business-impact and supply-risk evidence for Kraljic. Demand_History is the separate consumption fact for ABC/XYZ and ADI/CV2 item behavior; purchase orders are not used as demand."])
    read_me.append(["Decision guardrail", "Opportunity flags are review prompts, not booked savings. Validate addressability, contracts, continuity, criticality, and market conditions before action."])
    read_me.append([])
    read_me.append(["Sheet", "Use"])
    read_me.append(["Demand_Data", "Tidy historical demand records for upload and FPP3 analysis."])
    read_me.append(["Demand_Matrix", "One row per SKU with one column per historical month."])
    read_me.append(["Product_Master", "SKU attributes, approved cost, commercial risk context, legacy ABC/XYZ, default location, and supply type."])
    read_me.append(["DDMRP_Decoupling", "Approved synthetic decoupling design and DLT basis for each item-location."])
    read_me.append(["DDMRP_Master", "Approved synthetic buffer profile, ADU, factors, MOQ, order constraints, and spike settings."])
    read_me.append(["Inventory_Snapshot", "Current usable on-hand inventory at the planning snapshot."])
    read_me.append(["Open_Supply", "Open purchase/manufacturing supply with expected receipt and execution risk."])
    read_me.append(["Qualified_Demand", "Past-due, today, and qualified future spike demand used by the Net Flow Equation."])
    read_me.append(["DDMRP_Positions", "Calculator-ready daily position file with required DDMRP input columns."])
    read_me.append(["DDMRP_Recommendations", "Pre-calculated transparent synthetic order recommendations and execution review flags."])
    read_me.append(["DDMRP_Adjustments", "Current synthetic dynamic adjustment schedule. Replace with approved FPP3-derived proposals later."])
    read_me.append(["Spend_Lines", "Signed procurement transaction lines in EUR with supplier normalization, category, contract, FX, and classification lineage."])
    read_me.append(["Supplier_Master", "Normalized supplier and parent entities with commercial and operational risk fields."])
    read_me.append(["Contracts", "Synthetic active contract register used to distinguish on-contract spend from leakage."])
    read_me.append(["Category_Taxonomy", "Approved three-level procurement category mapping and ownership."])
    read_me.append(["Category_Risk", "Weighted Kraljic business-impact and supply-risk evidence at category grain."])
    read_me.append(["Spend_Control", "Signed AP ledger control total and reconciliation status."])
    read_me.append(["Demand_History", "Consumption history at part-location-period grain with approved unit cost for modern item segmentation."])
    read_me.append(["Data_Dictionary", "Field definitions for Demand_Data."])
    read_me.append(["DDMRP_Data_Dictionary", "Field definitions for the DDMRP sheets."])
    read_me.append(["Spend_Data_Dictionary", "Field definitions and method notes for procurement and portfolio analysis."])
    for row in read_me.iter_rows(min_row=3, max_row=read_me.max_row, min_col=1, max_col=2):
        row[0].font = HEADER_FONT
        row[0].fill = SUBHEADER_FILL
        row[0].border = Border(bottom=THIN_LINE)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    read_me.column_dimensions["A"].width = 25
    read_me.column_dimensions["B"].width = 115
    read_me.sheet_view.showGridLines = False

    master = workbook.create_sheet("Product_Master")
    master.append([
        "Product_Group_Code", "Product_Group", "SKU", "SKU_Description", "Demand_Profile", "Unit_Price_EUR",
        "Lead_Time_Days", "ABC_Class", "XYZ_Class", "Default_Location", "Supply_Type", "Unit_Cost_EUR",
        "Operational_Criticality", "Criticality_Score", "Lifecycle_Status", "Qualified_Source_Count", "Single_Source",
        "Substitute_Count", "Revenue_At_Risk_EUR", "Spend_Category_Code",
    ])
    demand = workbook.create_sheet("Demand_Data")
    demand.append([
        "Month", "Product_Group_Code", "Product_Group", "SKU", "SKU_Description", "Demand_Units",
        "Unit_Price_EUR", "Demand_Value_EUR", "Lead_Time_Days", "Demand_Profile",
    ])
    matrix = workbook.create_sheet("Demand_Matrix")
    matrix.append([
        "Product_Group_Code", "Product_Group", "SKU", "SKU_Description", "Demand_Profile",
        *[month.strftime("%Y-%m") for month in MONTHS],
    ])

    all_skus: list[dict[str, str | float | int]] = []
    histories: dict[str, list[int]] = {}
    for group in GROUPS:
        for sku_index in range(8):
            sku_data = make_sku(group, sku_index)
            all_skus.append(sku_data)
            master.append([
                sku_data["product_group_code"], sku_data["product_group"], sku_data["sku"], sku_data["description"],
                sku_data["profile"], sku_data["unit_price"], sku_data["lead_time"], sku_data["abc_class"], sku_data["xyz_class"],
                LOCATION, sku_data["supply_type"], sku_data["unit_cost"], sku_data["criticality"], sku_data["criticality_score"],
                sku_data["lifecycle"], sku_data["source_count"], sku_data["single_source"], sku_data["substitute_count"],
                sku_data["revenue_at_risk"], sku_data["spend_category_code"],
            ])

            history: list[int] = []
            for month_index, month in enumerate(MONTHS):
                demand_units = calculate_demand(group, sku_index, month_index, rng)
                history.append(demand_units)
                demand.append([
                    month, sku_data["product_group_code"], sku_data["product_group"], sku_data["sku"], sku_data["description"],
                    demand_units, sku_data["unit_price"], round(demand_units * float(sku_data["unit_price"]), 2),
                    sku_data["lead_time"], sku_data["profile"],
                ])
            histories[str(sku_data["sku"])] = history
            matrix.append([
                sku_data["product_group_code"], sku_data["product_group"], sku_data["sku"], sku_data["description"],
                sku_data["profile"], *history,
            ])

    ddmrp = build_ddmrp_rows(all_skus, histories, rng)
    decoupling = workbook.create_sheet("DDMRP_Decoupling")
    decoupling.append([
        "Snapshot_Date", "Location", "SKU", "SKU_Description", "Supply_Type", "Decoupling_Status", "DLT_Days", "DLT_Basis",
        "Upstream_Source", "Decoupling_Rationale", "Planner_Owner", "Approved_On", "Next_Review_Date", "Approval_Status",
    ])
    for row in ddmrp["decoupling"]:
        decoupling.append(row)

    ddmrp_master = workbook.create_sheet("DDMRP_Master")
    ddmrp_master.append([
        "Snapshot_Date", "Location", "SKU", "Product_Group", "SKU_Description", "Buffer_Profile", "Supply_Type", "Baseline_ADU",
        "ADU_Calculation_Method", "DLT_Days", "Lead_Time_Factor", "Variability_Factor", "MOQ", "Order_Cycle_Days",
        "Order_Multiple", "Spike_Horizon_Days", "Spike_Threshold_Units", "Demand_Adjustment_Factor", "Adjustment_Status",
        "Planner_Owner", "Setting_Approval_Status", "Next_Review_Date",
    ])
    for row in ddmrp["master"]:
        ddmrp_master.append(row)

    inventory = workbook.create_sheet("Inventory_Snapshot")
    inventory.append([
        "Snapshot_Date", "Source_Snapshot_At", "Location", "SKU", "SKU_Description", "On_Hand_Units", "Quality_Hold_Units",
        "Usable_On_Hand_Units", "Inventory_Status", "Data_Quality_Status",
    ])
    for row in ddmrp["inventory"]:
        inventory.append(row)

    open_supply = workbook.create_sheet("Open_Supply")
    open_supply.append([
        "Supply_Order_ID", "Location", "SKU", "Supply_Type", "Open_Supply_Units", "Expected_Receipt_Date", "Order_Status",
        "Eligible_for_NFP", "Supplier_or_Work_Center", "Synchronization_Note",
    ])
    for row in ddmrp["supply"]:
        open_supply.append(row)

    qualified_demand = workbook.create_sheet("Qualified_Demand")
    qualified_demand.append([
        "Demand_Reference", "Location", "SKU", "Due_Date", "Demand_Type", "Demand_Units", "Qualified_for_NFP",
        "Qualification_Reason", "Spike_Threshold_Units",
    ])
    for row in ddmrp["qualified_demand"]:
        qualified_demand.append(row)

    positions = workbook.create_sheet("DDMRP_Positions")
    positions.append([
        "SKU", "Location", "SKU_Description", "Product_Group", "ADU", "DLT_Days", "Lead_Time_Factor", "Variability_Factor",
        "MOQ", "Order_Cycle_Days", "On_Hand", "Open_Supply", "Past_Due_Demand", "Today_Demand", "Qualified_Spike_Demand",
        "Order_Multiple", "Demand_Adjustment_Factor", "Source_Snapshot_At",
    ])
    for row in ddmrp["positions"]:
        positions.append(row)

    recommendations = workbook.create_sheet("DDMRP_Recommendations")
    recommendations.append([
        "Planning_Date", "Location", "SKU", "SKU_Description", "Product_Group", "Buffer_Profile", "ADU", "Demand_Adjustment_Factor",
        "Adjusted_ADU", "DLT_Days", "Lead_Time_Factor", "Variability_Factor", "MOQ", "Order_Cycle_Days", "Order_Multiple",
        "Red_Base", "Red_Safety", "Red_Zone", "Yellow_Zone", "Green_Zone", "TOR", "TOY", "TOG", "On_Hand", "Open_Supply",
        "Past_Due_Demand", "Today_Demand", "Qualified_Spike_Demand", "Qualified_Demand", "Net_Flow_Position", "NFP_Percent_of_TOG",
        "NFP_Status", "Replenishment_Triggered", "Base_Recommended_Units", "Recommended_Order_Units", "Suggested_Availability_Date",
        "Planning_Action", "On_Hand_Gap_to_TOR", "Execution_Review", "Planner_Release_Status", "Data_Status",
    ])
    for row in ddmrp["recommendations"]:
        recommendations.append(row)

    adjustments = workbook.create_sheet("DDMRP_Adjustments")
    adjustments.append([
        "Location", "SKU", "Snapshot_Date", "Effective_From", "Effective_Through", "Demand_Adjustment_Factor", "Adjustment_Status",
        "Adjustment_Reason", "Forecast_Integration_Note", "Planner_Owner",
    ])
    for row in ddmrp["adjustments"]:
        adjustments.append(row)

    procurement = build_procurement_rows(all_skus, histories)

    spend_lines = workbook.create_sheet("Spend_Lines")
    spend_headers = [
        "Source_System", "Transaction_ID", "Line_ID", "Transaction_Date", "Supplier_ID", "Supplier_Name",
        "Supplier_Normalized_ID", "Supplier_Normalized_Name", "Supplier_Parent_ID", "Supplier_Parent_Name",
        "Category_L1", "Category_L2", "Category_L3", "Category_Code", "Description_Raw", "Description_Normalized",
        "Part_ID", "Organization", "Plant", "Cost_Center", "Buyer", "PO_ID", "PO_Date", "Invoice_ID",
        "Contract_ID", "On_Contract", "Preferred_Supplier", "Quantity", "UOM", "Quantity_Base_UOM",
        "Unit_Price_Base", "Spend_Base", "Base_Currency", "Spend_Original", "Original_Currency",
        "FX_Rate_to_EUR", "FX_Date", "Tax_Base", "Freight_Base", "Discount_Base", "Payment_Terms_Days",
        "Incoterm", "Direct_Indirect", "CapEx_OpEx", "Classification_Method", "Classification_Confidence",
        "Classification_Override", "Business_Impact_Score", "Supply_Risk_Score", "Data_Status",
    ]
    append_dict_rows(spend_lines, spend_headers, procurement["spend_lines"])

    supplier_master = workbook.create_sheet("Supplier_Master")
    supplier_headers = [
        "Supplier_ID", "Supplier_Normalized_ID", "Supplier_Normalized_Name", "Supplier_Parent_ID", "Supplier_Parent_Name",
        "Category_Code", "Category_L1", "Country", "Region", "Original_Currency", "Preferred_Supplier",
        "Approved_Status", "Supplier_Tier", "Financial_Risk_Score", "Quality_Risk_Score", "Delivery_Risk_Score",
        "ESG_Risk_Score", "Payment_Terms_Days", "Default_Incoterm", "Default_Contract_ID", "Risk_As_Of", "Risk_Owner",
        "Data_Status",
    ]
    append_dict_rows(supplier_master, supplier_headers, procurement["suppliers"])

    contracts = workbook.create_sheet("Contracts")
    contract_headers = [
        "Contract_ID", "Supplier_Normalized_ID", "Supplier_Normalized_Name", "Category_Code", "Category_L1",
        "Contract_Status", "Start_Date", "End_Date", "Contract_Owner", "Payment_Terms_Days", "Rebate_Percent",
        "Price_Adjustment_Clause", "Committed_Annual_Value_EUR", "Renewal_Action", "Data_Status",
    ]
    append_dict_rows(contracts, contract_headers, procurement["contracts"])

    category_taxonomy = workbook.create_sheet("Category_Taxonomy")
    taxonomy_headers = [
        "Category_Code", "Category_L1", "Category_L2", "Category_L3", "Direct_Indirect", "Category_Owner",
        "Taxonomy_Status", "Effective_From", "Next_Review_Date",
    ]
    append_dict_rows(category_taxonomy, taxonomy_headers, procurement["taxonomy"])

    category_risk = workbook.create_sheet("Category_Risk")
    category_risk_headers = [
        "Review_Date", "Category_Code", "Category_L1", "Spend_Impact_Score", "Operational_Criticality_Score",
        "Revenue_at_Risk_Score", "Quality_Regulatory_Score", "Innovation_Impact_Score", "Source_Scarcity_Score",
        "Switching_Time_Score", "Lead_Time_Risk_Score", "Capacity_Risk_Score", "Country_Logistics_Risk_Score",
        "Financial_Risk_Score", "Quality_Risk_Score", "ESG_Risk_Score", "Business_Impact_Score", "Supply_Risk_Score",
        "Kraljic_Quadrant", "Qualified_Supplier_Count", "Switching_Time_Days", "Category_Risk_Owner",
        "Evidence_Status", "Next_Review_Date",
    ]
    append_dict_rows(category_risk, category_risk_headers, procurement["category_risk"])

    spend_control = workbook.create_sheet("Spend_Control")
    spend_control_headers = [
        "Control_ID", "Scope_Start", "Scope_End", "Organization", "Base_Currency", "Control_Net_Spend",
        "Analyzed_Net_Spend", "Difference", "Gross_Positive_Spend", "Credits", "Line_Count", "Control_Status",
        "Control_Note",
    ]
    append_dict_rows(spend_control, spend_control_headers, procurement["controls"])

    demand_history = workbook.create_sheet("Demand_History")
    demand_history_headers = [
        "Period", "Part_ID", "Part_Description", "Product_Group", "Location", "Demand_Units", "Unit_Cost_Base",
        "Base_Currency", "Criticality", "Criticality_Score", "Lifecycle_Status", "Lead_Time_Days", "Source_Count",
        "Data_Quality_Status",
    ]
    append_dict_rows(demand_history, demand_history_headers, procurement["demand_history"])

    dictionary = workbook.create_sheet("Data_Dictionary")
    dictionary.append(["Field", "Description", "Example"])
    for row in (
        ("Month", "First day of the monthly demand period. Use this field as the time axis.", "2026-06-01"),
        ("Product_Group_Code", "Stable code for the main product group.", "BRG"),
        ("Product_Group", "Main product group label.", "Industrial Bearings"),
        ("SKU", "Unique stock-keeping unit identifier.", "BRG-01"),
        ("SKU_Description", "Readable SKU description.", "Industrial Bearings Core"),
        ("Demand_Units", "Observed monthly demand quantity in units.", "428"),
        ("Unit_Price_EUR", "Synthetic price per unit in EUR.", "60.84"),
        ("Demand_Value_EUR", "Demand_Units multiplied by Unit_Price_EUR.", "26039.52"),
        ("Lead_Time_Days", "Nominal replenishment context; use approved DLT for DDMRP buffers.", "42"),
        ("Demand_Profile", "Designed behavior pattern to aid forecasting tests.", "Seasonal"),
    ):
        dictionary.append(row)

    ddmrp_dictionary = workbook.create_sheet("DDMRP_Data_Dictionary")
    ddmrp_dictionary.append(["Sheet", "Field", "Description"])
    for row in (
        ("DDMRP_Decoupling", "DLT_Days", "Approved decoupled lead time for the item-location; calculated from source/BOM design in a real implementation."),
        ("DDMRP_Master", "Baseline_ADU", "Trailing 12 complete months of demand divided by calendar days. Recalculate from approved demand/forecast policy."),
        ("DDMRP_Master", "Lead_Time_Factor", "Factor used in red base and green calculations; grouped by DLT category in this synthetic example."),
        ("DDMRP_Master", "Variability_Factor", "Factor used in red safety; grouped by synthetic XYZ class in this example."),
        ("DDMRP_Master", "Demand_Adjustment_Factor", "Time-effective dynamic factor applied to ADU. This workbook uses synthetic demo adjustments; replace with approved FPP3 proposals."),
        ("Inventory_Snapshot", "Usable_On_Hand_Units", "Inventory available for the buffered position at the snapshot. This is the On_Hand input to net flow."),
        ("Open_Supply", "Eligible_for_NFP", "Only eligible open supply is included in Net Flow Position. Receipt timing remains an execution risk."),
        ("Qualified_Demand", "Qualified_for_NFP", "Past due, today, and qualifying future spikes are included. A monthly forecast is not qualified demand."),
        ("DDMRP_Positions", "Open_Supply", "Aggregate of eligible open supply records for SKU + Location at the snapshot."),
        ("DDMRP_Positions", "Qualified_Spike_Demand", "Future daily demand above the spike threshold and inside the spike horizon."),
        ("DDMRP_Recommendations", "Net_Flow_Position", "On hand + open supply - qualified demand."),
        ("DDMRP_Recommendations", "Recommended_Order_Units", "TOG - NFP when NFP <= TOY, rounded up to the valid order multiple."),
        ("DDMRP_Recommendations", "Execution_Review", "On-hand and synchronization warning separate from the net-flow supply recommendation."),
        ("DDMRP_Adjustments", "Forecast_Integration_Note", "Forecast-derived adjustments require FPP3 validation and planner approval before operational use."),
    ):
        ddmrp_dictionary.append(row)

    spend_dictionary = workbook.create_sheet("Spend_Data_Dictionary")
    spend_dictionary.append(["Sheet", "Field_or_Method", "Description"])
    for row in (
        ("Spend_Lines", "Transaction_ID + Line_ID", "Unique source transaction-line key used for lineage and duplicate checks."),
        ("Spend_Lines", "Spend_Base", "Signed procurement line amount in EUR. Credits are negative and retained for net-spend reconciliation."),
        ("Spend_Lines", "Spend_Original / FX_Rate_to_EUR", "Original-currency value and dated conversion rate retained beside the EUR analytical amount."),
        ("Spend_Lines", "Supplier_Normalized_ID / Supplier_Parent_ID", "Normalized legal entity and ultimate parent used to avoid false fragmentation."),
        ("Spend_Lines", "On_Contract", "Explicit known contract status. Unknown must not be treated as off-contract in live data."),
        ("Spend_Control", "Control_Net_Spend", "Signed source-ledger control total. Analysis must reconcile before financial interpretation."),
        ("Pareto", "Gross positive spend", "Ranking basis for monotonic supplier/category Pareto. Signed net spend remains the finance control basis."),
        ("Category_Risk", "Business_Impact_Score", "Weighted score: spend 30%, operational criticality 30%, revenue at risk 20%, quality/regulatory 10%, innovation 10%."),
        ("Category_Risk", "Supply_Risk_Score", "Weighted score: scarcity 20%, switching 15%, lead time 15%, capacity 15%, country/logistics 10%, financial 10%, quality 10%, ESG 5%."),
        ("Category_Risk", "Kraljic_Quadrant", "Category-level portfolio classification using explicit impact and supply-risk evidence with 50/50 thresholds."),
        ("Demand_History", "Demand_Units", "Consumption demand kept separate from purchase orders; used for item behavior and usage-value segmentation."),
        ("Item segmentation", "ABC/XYZ", "Communication layer based on annualized usage value and demand CV; boundaries must be reported and sensitivity reviewed."),
        ("Item segmentation", "ADI/CV2", "Syntetos-Boylan demand behavior: smooth, erratic, intermittent, or lumpy using ADI 1.32 and CV2 0.49 boundaries."),
        ("Opportunity flags", "No savings claim", "Flags identify review queues only. Addressability, criticality, contracts, market conditions, and implementation cost require validation."),
    ):
        spend_dictionary.append(row)

    master_widths = {get_column_letter(index): 18 for index in range(1, master.max_column + 1)}
    master_widths.update({"A": 20, "B": 25, "C": 14, "D": 34, "E": 16, "J": 18, "M": 24, "P": 22, "S": 22})
    style_sheet(master, master_widths)
    style_sheet(demand, {"A": 14, "B": 20, "C": 25, "D": 14, "E": 34, "F": 14, "G": 16, "H": 18, "I": 16, "J": 16})
    matrix_widths = {"A": 20, "B": 25, "C": 14, "D": 34, "E": 16}
    matrix_widths.update({get_column_letter(column): 13 for column in range(6, 6 + len(MONTHS))})
    style_sheet(matrix, matrix_widths)
    style_sheet(decoupling, {"A": 14, "B": 14, "C": 14, "D": 34, "E": 12, "F": 18, "G": 12, "H": 42, "I": 34, "J": 50, "K": 18, "L": 14, "M": 16, "N": 24})
    style_sheet(ddmrp_master, {"A": 14, "B": 14, "C": 14, "D": 25, "E": 34, "F": 16, "G": 12, "H": 14, "I": 42, "J": 12, "K": 16, "L": 17, "M": 12, "N": 18, "O": 15, "P": 18, "Q": 20, "R": 24, "S": 22, "T": 18, "U": 25, "V": 16})
    style_sheet(inventory, {"A": 14, "B": 24, "C": 14, "D": 14, "E": 34, "F": 18, "G": 18, "H": 22, "I": 16, "J": 34})
    style_sheet(open_supply, {"A": 20, "B": 14, "C": 14, "D": 14, "E": 20, "F": 22, "G": 14, "H": 18, "I": 24, "J": 42})
    style_sheet(qualified_demand, {"A": 20, "B": 14, "C": 14, "D": 14, "E": 18, "F": 14, "G": 18, "H": 48, "I": 22})
    style_sheet(positions, {"A": 14, "B": 14, "C": 34, "D": 25, "E": 12, "F": 12, "G": 16, "H": 17, "I": 12, "J": 18, "K": 14, "L": 16, "M": 18, "N": 14, "O": 24, "P": 15, "Q": 24, "R": 24})
    style_sheet(recommendations, {get_column_letter(index): 18 for index in range(1, 42)})
    recommendations.column_dimensions["D"].width = 34
    recommendations.column_dimensions["E"].width = 25
    recommendations.column_dimensions["AG"].width = 54
    recommendations.column_dimensions["AM"].width = 44
    recommendations.column_dimensions["AO"].width = 30
    style_sheet(adjustments, {"A": 14, "B": 14, "C": 14, "D": 16, "E": 18, "F": 24, "G": 24, "H": 46, "I": 70, "J": 18})
    spend_widths = {get_column_letter(index): 18 for index in range(1, spend_lines.max_column + 1)}
    spend_widths.update({"A": 20, "B": 24, "F": 34, "H": 34, "J": 34, "K": 25, "L": 28, "O": 38, "P": 38, "R": 34, "U": 20, "AS": 50, "AX": 38})
    style_sheet(spend_lines, spend_widths)
    supplier_widths = {get_column_letter(index): 18 for index in range(1, supplier_master.max_column + 1)}
    supplier_widths.update({"C": 34, "E": 34, "G": 25, "V": 20, "W": 40})
    style_sheet(supplier_master, supplier_widths)
    contract_widths = {get_column_letter(index): 18 for index in range(1, contracts.max_column + 1)}
    contract_widths.update({"C": 34, "E": 25, "L": 48, "N": 28, "O": 40})
    style_sheet(contracts, contract_widths)
    style_sheet(category_taxonomy, {"A": 18, "B": 25, "C": 28, "D": 25, "E": 18, "F": 20, "G": 28, "H": 16, "I": 18})
    risk_widths = {get_column_letter(index): 18 for index in range(1, category_risk.max_column + 1)}
    risk_widths.update({"C": 25, "S": 20, "V": 22, "W": 38})
    style_sheet(category_risk, risk_widths)
    style_sheet(spend_control, {"A": 34, "B": 16, "C": 16, "D": 36, "E": 16, "F": 22, "G": 22, "H": 16, "I": 24, "J": 18, "K": 16, "L": 18, "M": 56})
    style_sheet(demand_history, {"A": 14, "B": 14, "C": 38, "D": 25, "E": 14, "F": 16, "G": 18, "H": 16, "I": 18, "J": 18, "K": 20, "L": 16, "M": 16, "N": 38})
    style_sheet(dictionary, {"A": 22, "B": 80, "C": 28})
    style_sheet(ddmrp_dictionary, {"A": 26, "B": 28, "C": 105})
    style_sheet(spend_dictionary, {"A": 24, "B": 38, "C": 110})

    format_columns(master, integers=("Lead_Time_Days", "Criticality_Score", "Qualified_Source_Count", "Substitute_Count"))
    format_columns(demand, dates=("Month",), integers=("Demand_Units", "Lead_Time_Days"))
    for row in demand.iter_rows(min_row=2, max_row=demand.max_row, min_col=1, max_col=10):
        row[6].number_format = '#,##0.00 "EUR"'
        row[7].number_format = '#,##0.00 "EUR"'
    for row in master.iter_rows(min_row=2, max_row=master.max_row, min_col=1, max_col=master.max_column):
        row[5].number_format = '#,##0.00 "EUR"'
    format_currency_columns(master, ("Unit_Cost_EUR", "Revenue_At_Risk_EUR"))
    for row in matrix.iter_rows(min_row=2, max_row=matrix.max_row, min_col=6, max_col=matrix.max_column):
        for cell in row:
            cell.number_format = "#,##0"
    format_columns(decoupling, dates=("Snapshot_Date", "Approved_On", "Next_Review_Date"), integers=("DLT_Days",))
    format_columns(ddmrp_master, dates=("Snapshot_Date", "Next_Review_Date"), integers=("DLT_Days", "MOQ", "Order_Cycle_Days", "Order_Multiple", "Spike_Horizon_Days", "Spike_Threshold_Units"), decimals=("Baseline_ADU", "Lead_Time_Factor", "Variability_Factor", "Demand_Adjustment_Factor"))
    format_columns(inventory, dates=("Snapshot_Date",), integers=("On_Hand_Units", "Quality_Hold_Units", "Usable_On_Hand_Units"))
    format_columns(open_supply, dates=("Expected_Receipt_Date",), integers=("Open_Supply_Units",))
    format_columns(qualified_demand, dates=("Due_Date",), integers=("Demand_Units", "Spike_Threshold_Units"))
    format_columns(positions, integers=("DLT_Days", "MOQ", "Order_Cycle_Days", "On_Hand", "Open_Supply", "Past_Due_Demand", "Today_Demand", "Qualified_Spike_Demand", "Order_Multiple"), decimals=("ADU", "Lead_Time_Factor", "Variability_Factor", "Demand_Adjustment_Factor"))
    format_columns(recommendations, dates=("Planning_Date", "Suggested_Availability_Date"), integers=("DLT_Days", "MOQ", "Order_Cycle_Days", "Order_Multiple", "On_Hand", "Open_Supply", "Past_Due_Demand", "Today_Demand", "Qualified_Spike_Demand", "Qualified_Demand", "Net_Flow_Position", "Base_Recommended_Units", "Recommended_Order_Units"), decimals=("ADU", "Demand_Adjustment_Factor", "Adjusted_ADU", "Lead_Time_Factor", "Variability_Factor", "Red_Base", "Red_Safety", "Red_Zone", "Yellow_Zone", "Green_Zone", "TOR", "TOY", "TOG", "NFP_Percent_of_TOG", "On_Hand_Gap_to_TOR"))
    format_columns(adjustments, dates=("Snapshot_Date", "Effective_From", "Effective_Through"), decimals=("Demand_Adjustment_Factor",))
    format_columns(
        spend_lines,
        dates=("Transaction_Date", "PO_Date", "FX_Date"),
        integers=("Quantity", "Quantity_Base_UOM", "Payment_Terms_Days"),
        decimals=("Unit_Price_Base", "Spend_Base", "Spend_Original", "FX_Rate_to_EUR", "Tax_Base", "Freight_Base", "Discount_Base", "Classification_Confidence", "Business_Impact_Score", "Supply_Risk_Score"),
    )
    format_currency_columns(spend_lines, ("Unit_Price_Base", "Spend_Base", "Tax_Base", "Freight_Base", "Discount_Base"))
    format_columns(
        supplier_master,
        dates=("Risk_As_Of",),
        integers=("Financial_Risk_Score", "Quality_Risk_Score", "Delivery_Risk_Score", "ESG_Risk_Score", "Payment_Terms_Days"),
    )
    format_columns(contracts, dates=("Start_Date", "End_Date"), integers=("Payment_Terms_Days",), decimals=("Rebate_Percent", "Committed_Annual_Value_EUR"))
    format_currency_columns(contracts, ("Committed_Annual_Value_EUR",))
    format_columns(category_taxonomy, dates=("Effective_From", "Next_Review_Date"))
    format_columns(
        category_risk,
        dates=("Review_Date", "Next_Review_Date"),
        integers=("Spend_Impact_Score", "Operational_Criticality_Score", "Revenue_at_Risk_Score", "Quality_Regulatory_Score", "Innovation_Impact_Score", "Source_Scarcity_Score", "Switching_Time_Score", "Lead_Time_Risk_Score", "Capacity_Risk_Score", "Country_Logistics_Risk_Score", "Financial_Risk_Score", "Quality_Risk_Score", "ESG_Risk_Score", "Qualified_Supplier_Count", "Switching_Time_Days"),
        decimals=("Business_Impact_Score", "Supply_Risk_Score"),
    )
    format_columns(spend_control, dates=("Scope_Start", "Scope_End"), integers=("Line_Count",), decimals=("Control_Net_Spend", "Analyzed_Net_Spend", "Difference", "Gross_Positive_Spend", "Credits"))
    format_currency_columns(spend_control, ("Control_Net_Spend", "Analyzed_Net_Spend", "Difference", "Gross_Positive_Spend", "Credits"))
    format_columns(demand_history, dates=("Period",), integers=("Demand_Units", "Criticality_Score", "Lead_Time_Days", "Source_Count"), decimals=("Unit_Cost_Base",))
    format_currency_columns(demand_history, ("Unit_Cost_Base",))

    add_table(master, "ProductMaster")
    add_table(demand, "DemandData")
    add_table(matrix, "DemandMatrix")
    add_table(decoupling, "DDMRPDecoupling")
    add_table(ddmrp_master, "DDMRPMaster")
    add_table(inventory, "InventorySnapshot")
    add_table(open_supply, "OpenSupply")
    add_table(qualified_demand, "QualifiedDemand")
    add_table(positions, "DDMRPPositions")
    add_table(recommendations, "DDMRPRecommendations")
    add_table(adjustments, "DDMRPAdjustments")
    add_table(spend_lines, "SpendLines")
    add_table(supplier_master, "SupplierMaster")
    add_table(contracts, "Contracts")
    add_table(category_taxonomy, "CategoryTaxonomy")
    add_table(category_risk, "CategoryRisk")
    add_table(spend_control, "SpendControl")
    add_table(demand_history, "DemandHistory")
    add_table(dictionary, "DataDictionary")
    add_table(ddmrp_dictionary, "DDMRPDataDictionary")
    add_table(spend_dictionary, "SpendDataDictionary")

    workbook.save(OUTPUT_PATH)
    print(
        f"Created {OUTPUT_PATH} with {len(all_skus)} SKUs, {demand.max_row - 1} demand records, "
        f"{recommendations.max_row - 1} DDMRP recommendations, and {spend_lines.max_row - 1} signed spend lines."
    )


if __name__ == "__main__":
    build_workbook()
