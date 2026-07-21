"""Build the self-contained Demand Genie version 2 dashboard."""

from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = ROOT / "src" / "dashboard-v2.template.html"
STYLE_PATH = ROOT / "src" / "dashboard-v2.css"
SCRIPT_PATH = ROOT / "src" / "dashboard-v2.js"
SHEETJS_PATH = ROOT / "vendor" / "xlsx.full.min.js"
WORKBOOK_PATH = ROOT / "data" / "Demand_Genie_Synthetic_Demand_History.xlsx"
FORECAST_DIR = ROOT / "data" / "forecast"
OUTPUT_PATH = ROOT / "dashboard-v2.html"

SHEET_COLUMNS = {
    "Demand_Data": (
        "Month",
        "Product_Group_Code",
        "Product_Group",
        "SKU",
        "SKU_Description",
        "Demand_Units",
        "Demand_Profile",
    ),
    "Product_Master": None,
    "DDMRP_Decoupling": None,
    "DDMRP_Master": None,
    "Inventory_Snapshot": None,
    "Open_Supply": None,
    "Qualified_Demand": None,
    "DDMRP_Positions": None,
    "DDMRP_Recommendations": None,
    "DDMRP_Adjustments": None,
}

FORECAST_FILES = {
    "Forecast_Results": "forecast_results.csv",
    "Model_Selection": "model_selection.csv",
    "Forecast_Exceptions": "forecast_exceptions.csv",
    "Forecast_Run_Metadata": "run_metadata.csv",
}

NUMBER_PATTERN = re.compile(r"^-?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?$")


def json_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def worksheet_records(workbook, sheet_name: str, selected_columns: tuple[str, ...] | None) -> list[dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        return []
    rows = workbook[sheet_name].iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    columns = list(selected_columns) if selected_columns else headers
    indices = {header: index for index, header in enumerate(headers)}
    return [
        {column: json_value(row[indices[column]]) if column in indices and indices[column] < len(row) else None for column in columns}
        for row in rows
    ]


def csv_value(value: str) -> object:
    if value == "":
        return None
    if value == "TRUE":
        return True
    if value == "FALSE":
        return False
    if NUMBER_PATTERN.match(value):
        number = float(value)
        return int(number) if number.is_integer() else number
    return value


def csv_records(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as source:
        return [{key: csv_value(value) for key, value in row.items()} for row in csv.DictReader(source)]


def build_sample_data() -> dict[str, object]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheets = {
        sheet_name: worksheet_records(workbook, sheet_name, columns)
        for sheet_name, columns in SHEET_COLUMNS.items()
    }
    sheets.update(
        {
            sheet_name: csv_records(FORECAST_DIR / filename)
            for sheet_name, filename in FORECAST_FILES.items()
        }
    )
    return {
        "fileName": WORKBOOK_PATH.name,
        "sourceType": "Embedded synthetic demo",
        "forecastEngine": "R fpp3 1.0.2",
        "sheets": sheets,
    }


def main() -> None:
    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    styles = STYLE_PATH.read_text(encoding="utf-8")
    application = SCRIPT_PATH.read_text(encoding="utf-8")
    sheetjs = SHEETJS_PATH.read_text(encoding="utf-8")
    sample_json = json.dumps(build_sample_data(), ensure_ascii=True, separators=(",", ":")).replace("<", "\\u003c")

    output = template
    output = output.replace("/* __DASHBOARD_STYLES__ */", styles)
    output = output.replace("/* __SHEETJS_LIBRARY__ */", sheetjs)
    output = output.replace("__SAMPLE_DATA__", sample_json)
    output = output.replace("/* __DASHBOARD_APPLICATION__ */", application)
    if "__DASHBOARD_" in output or "__SAMPLE_DATA__" in output:
        raise RuntimeError("One or more dashboard build placeholders were not replaced")

    OUTPUT_PATH.write_text(output, encoding="utf-8")
    print(f"Built {OUTPUT_PATH} ({OUTPUT_PATH.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
